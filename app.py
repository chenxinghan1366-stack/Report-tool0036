import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import uuid
import sqlite3
import os
import zipfile
import shutil
import json
import hashlib
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText

# ========== 配置 ==========
DB_PATH = os.path.join(os.path.dirname(__file__), "app_data.db")
BACKUP_DIR = os.path.join(os.path.dirname(__file__), "backup")
TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "custom_templates")
ATTACHMENTS_DIR = os.path.join(os.path.dirname(__file__), "attachments")

for d in [BACKUP_DIR, TEMPLATES_DIR, ATTACHMENTS_DIR]:
    if not os.path.exists(d):
        os.makedirs(d)

def backup_database():
    if not os.path.exists(DB_PATH):
        return
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"data_backup_{timestamp}.db")
    try:
        shutil.copy2(DB_PATH, backup_path)
        backups = sorted([f for f in os.listdir(BACKUP_DIR) if f.startswith("data_backup_")])
        if len(backups) > 30:
            for f in backups[:-30]:
                os.remove(os.path.join(BACKUP_DIR, f))
        return backup_path
    except Exception as e:
        st.error(f"备份失败：{e}")
        return None

def get_backup_list():
    if not os.path.exists(BACKUP_DIR):
        return []
    return sorted([f for f in os.listdir(BACKUP_DIR) if f.startswith("data_backup_")], reverse=True)

def restore_backup(backup_file):
    backup_path = os.path.join(BACKUP_DIR, backup_file)
    if not os.path.exists(backup_path):
        return False
    try:
        backup_database()
        shutil.copy2(backup_path, DB_PATH)
        return True
    except Exception as e:
        st.error(f"恢复失败：{e}")
        return False

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS companies (
        id TEXT PRIMARY KEY, company_name TEXT, province TEXT, city TEXT, district TEXT, tax_id TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS templates (
        id TEXT PRIMARY KEY, province TEXT, city TEXT, district TEXT, report_type TEXT,
        template_name TEXT, template_version TEXT, source_url TEXT, source_authority TEXT,
        publish_date TEXT, required_fields TEXT, status TEXT, file_hash TEXT, file_type TEXT,
        is_custom BOOLEAN DEFAULT 0, field_mapping_source TEXT, use_count INTEGER DEFAULT 0
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS custom_templates (
        id TEXT PRIMARY KEY, name TEXT, file_data BLOB, field_mapping TEXT,
        sheet_name TEXT, created_at TEXT, updated_at TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS rules (
        id TEXT PRIMARY KEY, city TEXT, province TEXT, unit_social REAL, personal_social REAL,
        unit_fund REAL, personal_fund REAL, social_min REAL, social_max REAL,
        fund_min REAL, fund_max REAL, source_quote TEXT, is_default BOOLEAN DEFAULT 0,
        rule_version TEXT, effective_date TEXT,
        source_url TEXT, source_title TEXT, source_publish_date TEXT,
        collected_at TEXT, applicable_region TEXT, official_channel TEXT, notes TEXT,
        use_count INTEGER DEFAULT 0
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS export_history (
        id TEXT PRIMARY KEY, company_id TEXT, template_id TEXT, company_name TEXT,
        city TEXT, province TEXT, report_type TEXT, period_type TEXT, generated_at TEXT,
        review_status TEXT, reviewer TEXT, reviewed_at TEXT, file_name TEXT, file_data BLOB,
        data_source TEXT, month_used TEXT, year_used TEXT, custom_period TEXT,
        batch_id TEXT, job_name TEXT, field_mapping TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS source_registry (
        id TEXT PRIMARY KEY, authority_type TEXT, province TEXT, city TEXT, district TEXT,
        authority_name TEXT, official_site_name TEXT, source_url TEXT, source_level TEXT,
        source_section TEXT, is_official BOOLEAN, crawl_allowed BOOLEAN, last_checked TEXT,
        status TEXT, notes TEXT, document_name TEXT, document_version TEXT, publish_year TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS job_batches (
        id TEXT PRIMARY KEY, batch_name TEXT, created_at TEXT, status TEXT,
        total_companies INTEGER, total_reports INTEGER, review_status TEXT,
        parameters TEXT, created_by TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS job_details (
        id TEXT PRIMARY KEY, batch_id TEXT, company_id TEXT, company_name TEXT,
        city TEXT, province TEXT, report_type TEXT, period_type TEXT,
        status TEXT, error_message TEXT, file_name TEXT, file_data BLOB,
        generated_at TEXT, rule_source TEXT, data_source TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS verification_status (
        id TEXT PRIMARY KEY, batch_id TEXT,
        source_verified BOOLEAN DEFAULT 0, template_verified BOOLEAN DEFAULT 0,
        rule_verified BOOLEAN DEFAULT 0, data_verified BOOLEAN DEFAULT 0,
        reviewer_name TEXT, verified_at TEXT, notes TEXT, export_type TEXT DEFAULT '验证版'
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS rule_history (
        id TEXT PRIMARY KEY, rule_id TEXT, action_type TEXT,
        old_values TEXT, new_values TEXT, changed_by TEXT, changed_at TEXT, notes TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS template_history (
        id TEXT PRIMARY KEY, template_id TEXT, action_type TEXT,
        old_values TEXT, new_values TEXT, changed_by TEXT, changed_at TEXT, notes TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS audit_log (
        id TEXT PRIMARY KEY, user_name TEXT, action_type TEXT,
        target_type TEXT, target_id TEXT, details TEXT, ip_address TEXT, created_at TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS column_mappings (
        id TEXT PRIMARY KEY, user_id TEXT, sheet_name TEXT,
        city_col TEXT, company_col TEXT, district_col TEXT,
        created_at TEXT, updated_at TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS data_versions (
        id TEXT PRIMARY KEY, version_number INTEGER, created_at TEXT,
        record_count INTEGER, file_name TEXT, file_data BLOB, notes TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS email_config (
        id TEXT PRIMARY KEY, smtp_server TEXT, smtp_port INTEGER,
        sender_email TEXT, sender_password TEXT, recipient_email TEXT, updated_at TEXT
    )''')
    conn.commit()
    conn.close()

init_db()

def migrate_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("PRAGMA table_info(rules)")
    existing_cols = [col[1] for col in c.fetchall()]
    new_cols = ['source_url', 'source_title', 'source_publish_date', 'collected_at', 'applicable_region', 'official_channel', 'notes', 'use_count']
    for col in new_cols:
        if col not in existing_cols:
            if col == 'use_count':
                c.execute(f"ALTER TABLE rules ADD COLUMN {col} INTEGER DEFAULT 0")
            else:
                c.execute(f"ALTER TABLE rules ADD COLUMN {col} TEXT")
    c.execute("PRAGMA table_info(templates)")
    temp_cols = [col[1] for col in c.fetchall()]
    if 'use_count' not in temp_cols:
        c.execute("ALTER TABLE templates ADD COLUMN use_count INTEGER DEFAULT 0")
    tables_to_check = [
        ('templates', ['field_mapping_source']),
        ('export_history', ['batch_id', 'job_name', 'field_mapping']),
        ('source_registry', ['document_name', 'document_version', 'publish_year']),
    ]
    for table, cols in tables_to_check:
        c.execute(f"PRAGMA table_info({table})")
        existing = [col[1] for col in c.fetchall()]
        for col in cols:
            if col not in existing:
                c.execute(f"ALTER TABLE {table} ADD COLUMN {col} TEXT")
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='job_batches'")
    if not c.fetchone():
        c.execute('''CREATE TABLE job_batches (
            id TEXT PRIMARY KEY, batch_name TEXT, created_at TEXT, status TEXT,
            total_companies INTEGER, total_reports INTEGER, review_status TEXT,
            parameters TEXT, created_by TEXT
        )''')
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='job_details'")
    if not c.fetchone():
        c.execute('''CREATE TABLE job_details (
            id TEXT PRIMARY KEY, batch_id TEXT, company_id TEXT, company_name TEXT,
            city TEXT, province TEXT, report_type TEXT, period_type TEXT,
            status TEXT, error_message TEXT, file_name TEXT, file_data BLOB,
            generated_at TEXT, rule_source TEXT, data_source TEXT
        )''')
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='verification_status'")
    if not c.fetchone():
        c.execute('''CREATE TABLE verification_status (
            id TEXT PRIMARY KEY, batch_id TEXT,
            source_verified BOOLEAN DEFAULT 0, template_verified BOOLEAN DEFAULT 0,
            rule_verified BOOLEAN DEFAULT 0, data_verified BOOLEAN DEFAULT 0,
            reviewer_name TEXT, verified_at TEXT, notes TEXT, export_type TEXT DEFAULT '验证版'
        )''')
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='rule_history'")
    if not c.fetchone():
        c.execute('''CREATE TABLE rule_history (
            id TEXT PRIMARY KEY, rule_id TEXT, action_type TEXT,
            old_values TEXT, new_values TEXT, changed_by TEXT, changed_at TEXT, notes TEXT
        )''')
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='template_history'")
    if not c.fetchone():
        c.execute('''CREATE TABLE template_history (
            id TEXT PRIMARY KEY, template_id TEXT, action_type TEXT,
            old_values TEXT, new_values TEXT, changed_by TEXT, changed_at TEXT, notes TEXT
        )''')
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='audit_log'")
    if not c.fetchone():
        c.execute('''CREATE TABLE audit_log (
            id TEXT PRIMARY KEY, user_name TEXT, action_type TEXT,
            target_type TEXT, target_id TEXT, details TEXT, ip_address TEXT, created_at TEXT
        )''')
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='column_mappings'")
    if not c.fetchone():
        c.execute('''CREATE TABLE column_mappings (
            id TEXT PRIMARY KEY, user_id TEXT, sheet_name TEXT,
            city_col TEXT, company_col TEXT, district_col TEXT,
            created_at TEXT, updated_at TEXT
        )''')
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='data_versions'")
    if not c.fetchone():
        c.execute('''CREATE TABLE data_versions (
            id TEXT PRIMARY KEY, version_number INTEGER, created_at TEXT,
            record_count INTEGER, file_name TEXT, file_data BLOB, notes TEXT
        )''')
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='email_config'")
    if not c.fetchone():
        c.execute('''CREATE TABLE email_config (
            id TEXT PRIMARY KEY, smtp_server TEXT, smtp_port INTEGER,
            sender_email TEXT, sender_password TEXT, recipient_email TEXT, updated_at TEXT
        )''')
    conn.commit()
    conn.close()

migrate_db()

def dict_fetchall(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def safe_execute_query(query, params=None):
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        if params:
            c.execute(query, params)
        else:
            c.execute(query)
        rows = dict_fetchall(c)
        conn.close()
        return rows
    except sqlite3.OperationalError as e:
        if "no such table" in str(e):
            return []
        else:
            raise e

def load_companies():
    return safe_execute_query("SELECT * FROM companies")

def load_templates():
    return safe_execute_query("SELECT * FROM templates WHERE status='active' ORDER BY use_count DESC")

def load_custom_templates():
    return safe_execute_query("SELECT * FROM custom_templates")

def load_rules():
    return safe_execute_query("SELECT * FROM rules ORDER BY province, city")

def load_export_history():
    return safe_execute_query("SELECT * FROM export_history ORDER BY generated_at DESC")

def load_source_registry():
    return safe_execute_query("SELECT * FROM source_registry ORDER BY province, city")

def load_job_batches():
    return safe_execute_query("SELECT * FROM job_batches ORDER BY created_at DESC")

def load_job_details(batch_id):
    return safe_execute_query("SELECT * FROM job_details WHERE batch_id=? ORDER BY generated_at DESC", (batch_id,))

def load_verification_status(batch_id):
    rows = safe_execute_query("SELECT * FROM verification_status WHERE batch_id=?", (batch_id,))
    return rows[0] if rows else None

def load_rule_history(rule_id=None):
    if rule_id:
        return safe_execute_query("SELECT * FROM rule_history WHERE rule_id=? ORDER BY changed_at DESC", (rule_id,))
    return safe_execute_query("SELECT * FROM rule_history ORDER BY changed_at DESC")

def load_template_history(template_id=None):
    if template_id:
        return safe_execute_query("SELECT * FROM template_history WHERE template_id=? ORDER BY changed_at DESC", (template_id,))
    return safe_execute_query("SELECT * FROM template_history ORDER BY changed_at DESC")

def load_audit_log(limit=100):
    return safe_execute_query("SELECT * FROM audit_log ORDER BY created_at DESC LIMIT ?", (limit,))

def load_column_mappings(user_id="default"):
    rows = safe_execute_query("SELECT * FROM column_mappings WHERE user_id=?", (user_id,))
    return rows[0] if rows else None

def save_column_mapping(user_id, sheet_name, city_col, company_col, district_col):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    existing = c.execute("SELECT id FROM column_mappings WHERE user_id=?", (user_id,)).fetchone()
    if existing:
        c.execute('''UPDATE column_mappings 
            SET sheet_name=?, city_col=?, company_col=?, district_col=?, updated_at=?
            WHERE user_id=?''',
            (sheet_name, city_col, company_col, district_col, datetime.now().isoformat(), user_id))
    else:
        c.execute('''INSERT INTO column_mappings 
            (id, user_id, sheet_name, city_col, company_col, district_col, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?)''',
            (str(uuid.uuid4())[:8], user_id, sheet_name, city_col, company_col, district_col,
             datetime.now().isoformat(), datetime.now().isoformat()))
    conn.commit()
    conn.close()
    backup_database()

def load_data_versions(limit=10):
    return safe_execute_query("SELECT * FROM data_versions ORDER BY version_number DESC LIMIT ?", (limit,))

def save_data_version(record_count, file_name, file_data, notes=""):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    max_ver = c.execute("SELECT MAX(version_number) FROM data_versions").fetchone()[0]
    new_ver = (max_ver or 0) + 1
    c.execute('''INSERT INTO data_versions 
        (id, version_number, created_at, record_count, file_name, file_data, notes)
        VALUES (?,?,?,?,?,?,?)''',
        (str(uuid.uuid4())[:8], new_ver, datetime.now().isoformat(), 
         record_count, file_name, file_data, notes))
    conn.commit()
    conn.close()
    backup_database()

def load_email_config():
    rows = safe_execute_query("SELECT * FROM email_config LIMIT 1")
    return rows[0] if rows else None

def save_email_config(config):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    existing = c.execute("SELECT id FROM email_config LIMIT 1").fetchone()
    if existing:
        c.execute('''UPDATE email_config 
            SET smtp_server=?, smtp_port=?, sender_email=?, sender_password=?, recipient_email=?, updated_at=?
            WHERE id=?''',
            (config['smtp_server'], config['smtp_port'], config['sender_email'],
             config['sender_password'], config['recipient_email'], datetime.now().isoformat(), existing[0]))
    else:
        c.execute('''INSERT INTO email_config 
            (id, smtp_server, smtp_port, sender_email, sender_password, recipient_email, updated_at)
            VALUES (?,?,?,?,?,?,?)''',
            (str(uuid.uuid4())[:8], config['smtp_server'], config['smtp_port'],
             config['sender_email'], config['sender_password'], config['recipient_email'],
             datetime.now().isoformat()))
    conn.commit()
    conn.close()
    backup_database()

def send_email_with_attachment(recipient, subject, body, attachment_data, filename):
    config = load_email_config()
    if not config:
        return False, "请先配置邮件设置"
    try:
        msg = MIMEMultipart()
        msg['From'] = config['sender_email']
        msg['To'] = recipient or config['recipient_email']
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment_data)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)
        server = smtplib.SMTP(config['smtp_server'], config['smtp_port'])
        server.starttls()
        server.login(config['sender_email'], config['sender_password'])
        server.send_message(msg)
        server.quit()
        return True, "邮件发送成功"
    except Exception as e:
        return False, f"邮件发送失败: {str(e)}"

def log_audit(user_name, action_type, target_type, target_id, details, ip_address=""):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''INSERT INTO audit_log 
        (id, user_name, action_type, target_type, target_id, details, ip_address, created_at)
        VALUES (?,?,?,?,?,?,?,?)''',
        (str(uuid.uuid4())[:8], user_name, action_type, target_type, target_id, details, ip_address, datetime.now().isoformat()))
    conn.commit()
    conn.close()
    backup_database()

def log_rule_change(rule_id, action_type, old_values, new_values, changed_by, notes=""):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''INSERT INTO rule_history 
        (id, rule_id, action_type, old_values, new_values, changed_by, changed_at, notes)
        VALUES (?,?,?,?,?,?,?,?)''',
        (str(uuid.uuid4())[:8], rule_id, action_type, json.dumps(old_values), json.dumps(new_values), changed_by, datetime.now().isoformat(), notes))
    conn.commit()
    conn.close()
    backup_database()

def log_template_change(template_id, action_type, old_values, new_values, changed_by, notes=""):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''INSERT INTO template_history 
        (id, template_id, action_type, old_values, new_values, changed_by, changed_at, notes)
        VALUES (?,?,?,?,?,?,?,?)''',
        (str(uuid.uuid4())[:8], template_id, action_type, json.dumps(old_values), json.dumps(new_values), changed_by, datetime.now().isoformat(), notes))
    conn.commit()
    conn.close()
    backup_database()

def rollback_rule(history_id):
    history = safe_execute_query("SELECT * FROM rule_history WHERE id=?", (history_id,))
    if not history:
        return False
    h = history[0]
    old_vals = json.loads(h['old_values'])
    rules = load_rules()
    for r in rules:
        if r['id'] == h['rule_id']:
            r.update(old_vals)
            break
    save_rules(rules)
    log_audit("系统", "ROLLBACK", "rule", h['rule_id'], f"回滚到版本 {h['changed_at']}")
    return True

def rollback_template(history_id):
    history = safe_execute_query("SELECT * FROM template_history WHERE id=?", (history_id,))
    if not history:
        return False
    h = history[0]
    old_vals = json.loads(h['old_values'])
    templates = load_templates()
    for t in templates:
        if t['id'] == h['template_id']:
            t.update(old_vals)
            break
    save_template(t)
    log_audit("系统", "ROLLBACK", "template", h['template_id'], f"回滚到版本 {h['changed_at']}")
    return True

def save_verification_status(data):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''INSERT OR REPLACE INTO verification_status 
        (id, batch_id, source_verified, template_verified, rule_verified, data_verified,
         reviewer_name, verified_at, notes, export_type)
        VALUES (?,?,?,?,?,?,?,?,?,?)''',
        (data.get('id', str(uuid.uuid4())[:8]), data['batch_id'],
         data.get('source_verified', 0), data.get('template_verified', 0),
         data.get('rule_verified', 0), data.get('data_verified', 0),
         data.get('reviewer_name', ''), data.get('verified_at', datetime.now().isoformat()),
         data.get('notes', ''), data.get('export_type', '验证版')))
    conn.commit()
    conn.close()
    backup_database()

def update_verification_status(batch_id, field, value):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(f"UPDATE verification_status SET {field}=? WHERE batch_id=?", (value, batch_id))
    conn.commit()
    conn.close()
    backup_database()

def save_source_registry(sources):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM source_registry")
    for s in sources:
        c.execute('''INSERT OR REPLACE INTO source_registry 
            (id, authority_type, province, city, district, authority_name,
             official_site_name, source_url, source_level, source_section,
             is_official, crawl_allowed, last_checked, status, notes,
             document_name, document_version, publish_year)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (s.get('id', str(uuid.uuid4())[:8]), s.get('authority_type','tax'),
             s.get('province',''), s.get('city',''), s.get('district',''),
             s.get('authority_name',''), s.get('official_site_name',''),
             s.get('source_url',''), s.get('source_level',''), s.get('source_section',''),
             s.get('is_official',1), s.get('crawl_allowed',1),
             s.get('last_checked',''), s.get('status','active'), s.get('notes',''),
             s.get('document_name',''), s.get('document_version',''), s.get('publish_year','')))
    conn.commit()
    conn.close()
    backup_database()

def delete_source_by_id(source_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM source_registry WHERE id=?", (source_id,))
    conn.commit()
    conn.close()
    backup_database()

def save_companies(companies):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM companies")
    for comp in companies:
        c.execute('''INSERT INTO companies (id, company_name, province, city, district, tax_id)
            VALUES (?,?,?,?,?,?)''',
            (comp.get('id', str(uuid.uuid4())[:8]), comp['company_name'], comp['province'],
             comp['city'], comp.get('district',''), comp.get('tax_id','')))
    conn.commit()
    conn.close()
    backup_database()

def save_template(template):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id FROM templates WHERE source_url=? OR file_hash=?", 
              (template.get('source_url',''), template.get('file_hash','')))
    existing = c.fetchone()
    if existing:
        c.execute('''UPDATE templates SET 
            template_name=?, template_version=?, source_authority=?, publish_date=?,
            required_fields=?, status=?, file_type=?, is_custom=?, field_mapping_source=?, use_count=?
            WHERE id=?''',
            (template['template_name'], template.get('template_version','v1.0'),
             template.get('source_authority',''), template.get('publish_date',''),
             template.get('required_fields',''), template.get('status','active'),
             template.get('file_type',''), template.get('is_custom',0),
             template.get('field_mapping_source',''), template.get('use_count',0), existing[0]))
    else:
        c.execute('''INSERT INTO templates 
            (id, province, city, district, report_type, template_name, template_version,
             source_url, source_authority, publish_date, required_fields, status, 
             file_hash, file_type, is_custom, field_mapping_source, use_count)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (template['id'], template.get('province',''), template.get('city',''),
             template.get('district',''), template.get('report_type',''),
             template['template_name'], template.get('template_version','v1.0'),
             template.get('source_url',''), template.get('source_authority',''),
             template.get('publish_date',''), template.get('required_fields',''),
             template.get('status','active'), template.get('file_hash',''),
             template.get('file_type',''), template.get('is_custom',0),
             template.get('field_mapping_source',''), template.get('use_count',0)))
    conn.commit()
    conn.close()
    backup_database()

def increment_template_use(template_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("UPDATE templates SET use_count = use_count + 1 WHERE id=?", (template_id,))
    conn.commit()
    conn.close()
    backup_database()

def increment_rule_use(rule_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("UPDATE rules SET use_count = use_count + 1 WHERE id=?", (rule_id,))
    conn.commit()
    conn.close()
    backup_database()

def save_custom_template(template):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''INSERT OR REPLACE INTO custom_templates 
        (id, name, file_data, field_mapping, sheet_name, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?)''',
        (template['id'], template['name'], template['file_data'],
         json.dumps(template.get('field_mapping', {})),
         template.get('sheet_name', ''), template.get('created_at', datetime.now().isoformat()),
         datetime.now().isoformat()))
    conn.commit()
    conn.close()
    backup_database()

def delete_custom_template(template_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM custom_templates WHERE id=?", (template_id,))
    conn.commit()
    conn.close()
    backup_database()

def save_rules(rules):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM rules")
    for r in rules:
        c.execute('''INSERT INTO rules 
            (id, city, province, unit_social, personal_social, unit_fund, personal_fund,
             social_min, social_max, fund_min, fund_max, source_quote, is_default,
             rule_version, effective_date, source_url, source_title, source_publish_date,
             collected_at, applicable_region, official_channel, notes, use_count)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (r['id'], r['city'], r.get('province',''), r['unit_social'], r['personal_social'],
             r['unit_fund'], r['personal_fund'], r.get('social_min',0), r.get('social_max',999999),
             r.get('fund_min',0), r.get('fund_max',999999), r.get('source_quote',''),
             r.get('is_default',0), r.get('rule_version','v1.0'), r.get('effective_date',''),
             r.get('source_url',''), r.get('source_title',''), r.get('source_publish_date',''),
             r.get('collected_at', datetime.now().isoformat()), r.get('applicable_region',''),
             r.get('official_channel',''), r.get('notes',''), r.get('use_count',0)))
    conn.commit()
    conn.close()
    backup_database()

def save_export(record):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''INSERT OR REPLACE INTO export_history 
        (id, company_id, template_id, company_name, city, province, report_type, period_type,
         generated_at, review_status, reviewer, reviewed_at, file_name, file_data,
         data_source, month_used, year_used, custom_period, batch_id, job_name, field_mapping)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (record['id'], record.get('company_id',''), record.get('template_id',''),
         record['company_name'], record.get('city',''), record.get('province',''),
         record.get('report_type',''), record.get('period_type',''), record['generated_at'],
         record.get('review_status','pending'), record.get('reviewer',''), record.get('reviewed_at',''),
         record.get('file_name',''), record.get('file_data', None),
         record.get('data_source',''), record.get('month_used',''), record.get('year_used',''),
         record.get('custom_period',''), record.get('batch_id',''), record.get('job_name',''),
         record.get('field_mapping','')))
    conn.commit()
    conn.close()
    backup_database()

def save_batch_job(batch):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''INSERT OR REPLACE INTO job_batches 
        (id, batch_name, created_at, status, total_companies, total_reports, 
         review_status, parameters, created_by)
        VALUES (?,?,?,?,?,?,?,?,?)''',
        (batch['id'], batch['batch_name'], batch['created_at'], batch.get('status','pending'),
         batch.get('total_companies',0), batch.get('total_reports',0),
         batch.get('review_status','pending'), json.dumps(batch.get('parameters',{})),
         batch.get('created_by','系统')))
    conn.commit()
    conn.close()
    backup_database()

def save_job_details(details):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    for d in details:
        c.execute('''INSERT OR REPLACE INTO job_details 
            (id, batch_id, company_id, company_name, city, province, report_type, 
             period_type, status, error_message, file_name, file_data, generated_at,
             rule_source, data_source)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (d.get('id', str(uuid.uuid4())[:8]), d.get('batch_id',''), d.get('company_id',''),
             d.get('company_name',''), d.get('city',''), d.get('province',''),
             d.get('report_type',''), d.get('period_type',''), d.get('status','pending'),
             d.get('error_message',''), d.get('file_name',''), d.get('file_data', None),
             d.get('generated_at', datetime.now().isoformat()), d.get('rule_source',''),
             d.get('data_source','')))
    conn.commit()
    conn.close()
    backup_database()

def update_export_status(export_id, status, reviewer):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''UPDATE export_history 
        SET review_status=?, reviewer=?, reviewed_at=?
        WHERE id=?''',
        (status, reviewer, datetime.now().isoformat(), export_id))
    conn.commit()
    conn.close()
    backup_database()

def update_batch_status(batch_id, status, review_status=None):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    if review_status:
        c.execute('''UPDATE job_batches 
            SET status=?, review_status=?
            WHERE id=?''', (status, review_status, batch_id))
    else:
        c.execute('''UPDATE job_batches SET status=?
            WHERE id=?''', (status, batch_id))
    conn.commit()
    conn.close()
    backup_database()

def normalize_name(name):
    if not name:
        return name
    for suffix in ['省', '市', '区', '县', '自治区', '特别行政区', '自治州', '地区']:
        if name.endswith(suffix):
            name = name[:-len(suffix)]
    return name.strip()

# ========== 模板库预置 ==========
PRESET_TEMPLATES = [
    {'id': 'tpl_vat_general', 'province': '全国', 'city': '全国', 'district': '',
     'report_type': '增值税', 'template_name': '增值税一般纳税人申报表',
     'template_version': '2024.1', 'source_url': 'https://www.chinatax.gov.cn/',
     'source_authority': '国家税务总局', 'publish_date': '2024-01-01',
     'required_fields': '纳税人识别号,公司名称,销售额,进项税额,应纳税额',
     'status': 'active', 'file_hash': 'preset_vat', 'file_type': 'preset',
     'is_custom': 0, 'field_mapping_source': '自动映射', 'use_count': 0},
    {'id': 'tpl_social_general', 'province': '全国', 'city': '全国', 'district': '',
     'report_type': '社保', 'template_name': '社会保险费申报表',
     'template_version': '2024.1', 'source_url': 'https://www.mohrss.gov.cn/',
     'source_authority': '人力资源和社会保障部', 'publish_date': '2024-01-01',
     'required_fields': '单位名称,社保登记号,基数,单位金额,个人金额',
     'status': 'active', 'file_hash': 'preset_social', 'file_type': 'preset',
     'is_custom': 0, 'field_mapping_source': '自动映射', 'use_count': 0},
    {'id': 'tpl_fund_general', 'province': '全国', 'city': '全国', 'district': '',
     'report_type': '公积金', 'template_name': '住房公积金汇缴申报表',
     'template_version': '2024.1', 'source_url': 'https://www.mohrss.gov.cn/',
     'source_authority': '住房和城乡建设部', 'publish_date': '2024-01-01',
     'required_fields': '单位名称,公积金账号,基数,单位比例,个人比例,单位金额,个人金额',
     'status': 'active', 'file_hash': 'preset_fund', 'file_type': 'preset',
     'is_custom': 0, 'field_mapping_source': '自动映射', 'use_count': 0},
]

def init_preset_templates():
    existing = load_templates()
    existing_names = {t['template_name'] for t in existing}
    added = 0
    for pt in PRESET_TEMPLATES:
        if pt['template_name'] not in existing_names:
            save_template(pt)
            added += 1
    return added

init_preset_templates()

PROVINCE_DEFAULT_RULES = [
    {'city': '上海', 'province': '上海', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.07, 'personal_fund': 0.07, 'social_min': 7310, 'social_max': 36549,
     'fund_min': 2590, 'fund_max': 34188, 'source_quote': '沪人社规〔2024〕22号'},
    {'city': '北京', 'province': '北京', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 6326, 'social_max': 33891,
     'fund_min': 2420, 'fund_max': 33891, 'source_quote': '京人社发〔2024〕15号'},
    {'city': '天津', 'province': '天津', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.11, 'personal_fund': 0.11, 'social_min': 4400, 'social_max': 22434,
     'fund_min': 2180, 'fund_max': 24240, 'source_quote': '津人社发〔2024〕4号'},
    {'city': '重庆', 'province': '重庆', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 19784,
     'fund_min': 2100, 'fund_max': 24595, 'source_quote': '渝人社发〔2024〕5号'},
    {'city': '广州', 'province': '广东', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.10, 'personal_fund': 0.10, 'social_min': 4588, 'social_max': 22941,
     'fund_min': 2300, 'fund_max': 27960, 'source_quote': '穗人社发〔2024〕3号'},
    {'city': '深圳', 'province': '广东', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 2360, 'social_max': 22941,
     'fund_min': 2360, 'fund_max': 27927, 'source_quote': '深人社规〔2024〕3号'},
    {'city': '东莞', 'province': '广东', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.10, 'personal_fund': 0.10, 'social_min': 4588, 'social_max': 22941,
     'fund_min': 1900, 'fund_max': 25431, 'source_quote': '东人社发〔2024〕6号'},
    {'city': '佛山', 'province': '广东', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.10, 'personal_fund': 0.10, 'social_min': 4588, 'social_max': 22941,
     'fund_min': 1900, 'fund_max': 25431, 'source_quote': '佛人社发〔2024〕5号'},
    {'city': '南京', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.08, 'personal_fund': 0.08, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27841, 'source_quote': '宁人社发〔2024〕5号'},
    {'city': '苏州', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '苏人社发〔2024〕6号'},
    {'city': '杭州', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '杭人社发〔2024〕6号'},
    {'city': '宁波', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '甬人社发〔2024〕5号'},
    {'city': '成都', 'province': '四川', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4071, 'social_max': 20355,
     'fund_min': 2100, 'fund_max': 25401, 'source_quote': '成人社发〔2024〕7号'},
    {'city': '绵阳', 'province': '四川', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（建议核实）'},
    {'city': '武汉', 'province': '湖北', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4077, 'social_max': 20385,
     'fund_min': 2010, 'fund_max': 24114, 'source_quote': '武人社发〔2024〕4号'},
    {'city': '襄阳', 'province': '湖北', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（建议核实）'},
    {'city': '青岛', 'province': '山东', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3746, 'social_max': 18726,
     'fund_min': 2010, 'fund_max': 23496, 'source_quote': '青人社发〔2024〕4号'},
    {'city': '济南', 'province': '山东', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3746, 'social_max': 18726,
     'fund_min': 2010, 'fund_max': 23496, 'source_quote': '济人社发〔2024〕5号'},
    {'city': '西安', 'province': '陕西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.10, 'personal_fund': 0.10, 'social_min': 3957, 'social_max': 19784,
     'fund_min': 1950, 'fund_max': 23556, 'source_quote': '西人社发〔2024〕6号'},
    {'city': '沈阳', 'province': '辽宁', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4100, 'social_max': 20500,
     'fund_min': 2100, 'fund_max': 25200, 'source_quote': '沈人社发〔2024〕5号'},
    {'city': '大连', 'province': '辽宁', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4100, 'social_max': 20500,
     'fund_min': 2100, 'fund_max': 25200, 'source_quote': '大人社发〔2024〕4号'},
    {'city': '福州', 'province': '福建', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4100, 'social_max': 20500,
     'fund_min': 2100, 'fund_max': 25200, 'source_quote': '榕人社发〔2024〕5号'},
    {'city': '厦门', 'province': '福建', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4100, 'social_max': 20500,
     'fund_min': 2100, 'fund_max': 25200, 'source_quote': '厦人社发〔2024〕4号'},
    {'city': '石家庄', 'province': '河北', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3800, 'social_max': 19000,
     'fund_min': 1900, 'fund_max': 22800, 'source_quote': '石人社发〔2024〕5号'},
    {'city': '合肥', 'province': '安徽', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3900, 'social_max': 19500,
     'fund_min': 1950, 'fund_max': 23400, 'source_quote': '合人社发〔2024〕5号'},
    {'city': '南昌', 'province': '江西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3800, 'social_max': 19000,
     'fund_min': 1900, 'fund_max': 22800, 'source_quote': '洪人社发〔2024〕4号'},
    {'city': '太原', 'province': '山西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3700, 'social_max': 18500,
     'fund_min': 1850, 'fund_max': 22200, 'source_quote': '并人社发〔2024〕4号'},
    {'city': '长春', 'province': '吉林', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3700, 'social_max': 18500,
     'fund_min': 1850, 'fund_max': 22200, 'source_quote': '长人社发〔2024〕4号'},
    {'city': '哈尔滨', 'province': '黑龙江', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '哈人社发〔2024〕4号'},
    {'city': '昆明', 'province': '云南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3700, 'social_max': 18500,
     'fund_min': 1850, 'fund_max': 22200, 'source_quote': '昆人社发〔2024〕5号'},
    {'city': '贵阳', 'province': '贵州', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '筑人社发〔2024〕4号'},
    {'city': '兰州', 'province': '甘肃', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3500, 'social_max': 17500,
     'fund_min': 1750, 'fund_max': 21000, 'source_quote': '兰人社发〔2024〕4号'},
    {'city': '酒泉', 'province': '甘肃', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（甘肃）'},
    {'city': '张掖', 'province': '甘肃', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（甘肃）'},
    {'city': '西宁', 'province': '青海', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3400, 'social_max': 17000,
     'fund_min': 1700, 'fund_max': 20400, 'source_quote': '宁人社发〔2024〕4号'},
    {'city': '海东', 'province': '青海', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（青海）'},
    {'city': '格尔木', 'province': '青海', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（青海）'},
    {'city': '银川', 'province': '宁夏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3500, 'social_max': 17500,
     'fund_min': 1750, 'fund_max': 21000, 'source_quote': '银人社发〔2024〕4号'},
    {'city': '石嘴山', 'province': '宁夏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（宁夏）'},
    {'city': '吴忠', 'province': '宁夏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（宁夏）'},
    {'city': '呼和浩特', 'province': '内蒙古', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '呼人社发〔2024〕4号'},
    {'city': '包头', 'province': '内蒙古', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（内蒙古）'},
    {'city': '鄂尔多斯', 'province': '内蒙古', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（内蒙古）'},
    {'city': '乌鲁木齐', 'province': '新疆', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3500, 'social_max': 17500,
     'fund_min': 1750, 'fund_max': 21000, 'source_quote': '乌人社发〔2024〕4号'},
    {'city': '拉萨', 'province': '西藏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3300, 'social_max': 16500,
     'fund_min': 1650, 'fund_max': 19800, 'source_quote': '拉人社发〔2024〕4号'},
    {'city': '海口', 'province': '海南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3800, 'social_max': 19000,
     'fund_min': 1900, 'fund_max': 22800, 'source_quote': '海人社发〔2024〕4号'},
    {'city': '南宁', 'province': '广西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '南人社发〔2024〕4号'},
    {'city': '遵义', 'province': '贵州', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（贵州）'},
    {'city': '六盘水', 'province': '贵州', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（贵州）'},
    {'city': '曲靖', 'province': '云南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（云南）'},
    {'city': '玉溪', 'province': '云南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999,
     'fund_min': 0, 'fund_max': 999999, 'source_quote': '系统默认（云南）'},
]

def init_rules_from_default():
    existing = load_rules()
    if existing:
        existing_cities = {normalize_name(r['city']) for r in existing}
        added = 0
        for dr in PROVINCE_DEFAULT_RULES:
            norm_city = normalize_name(dr['city'])
            if norm_city not in existing_cities:
                new_rule = {
                    'id': str(uuid.uuid4())[:8],
                    'city': dr['city'],
                    'province': dr['province'],
                    'unit_social': dr['unit_social'],
                    'personal_social': dr['personal_social'],
                    'unit_fund': dr['unit_fund'],
                    'personal_fund': dr['personal_fund'],
                    'social_min': dr.get('social_min', 0),
                    'social_max': dr.get('social_max', 999999),
                    'fund_min': dr.get('fund_min', 0),
                    'fund_max': dr.get('fund_max', 999999),
                    'source_quote': dr.get('source_quote', '省份默认'),
                    'is_default': 1,
                    'rule_version': 'v1.0',
                    'effective_date': datetime.now().strftime('%Y-%m-%d'),
                    'source_url': dr.get('source_url', '#'),
                    'source_title': dr.get('source_title', '系统默认规则'),
                    'source_publish_date': datetime.now().strftime('%Y-%m-%d'),
                    'collected_at': datetime.now().isoformat(),
                    'applicable_region': dr['province'],
                    'official_channel': '系统默认',
                    'notes': '自动补全',
                    'use_count': 0
                }
                existing.append(new_rule)
                added += 1
        if added > 0:
            save_rules(existing)
        return
    all_rules = []
    for dr in PROVINCE_DEFAULT_RULES:
        all_rules.append({
            'id': str(uuid.uuid4())[:8],
            'city': dr['city'],
            'province': dr.get('province', dr['city']),
            'unit_social': dr['unit_social'],
            'personal_social': dr['personal_social'],
            'unit_fund': dr['unit_fund'],
            'personal_fund': dr['personal_fund'],
            'social_min': dr.get('social_min', 0),
            'social_max': dr.get('social_max', 999999),
            'fund_min': dr.get('fund_min', 0),
            'fund_max': dr.get('fund_max', 999999),
            'source_quote': dr.get('source_quote', '省份默认'),
            'is_default': 1,
            'rule_version': 'v1.0',
            'effective_date': datetime.now().strftime('%Y-%m-%d'),
            'source_url': dr.get('source_url', '#'),
            'source_title': dr.get('source_title', '系统默认规则'),
            'source_publish_date': datetime.now().strftime('%Y-%m-%d'),
            'collected_at': datetime.now().isoformat(),
            'applicable_region': dr.get('province', dr['city']),
            'official_channel': '系统默认',
            'notes': '初始化',
            'use_count': 0
        })
    save_rules(all_rules)

init_rules_from_default()

def match_template_with_details(province, city, district, report_type):
    templates = load_templates()
    if not templates:
        return None, None, []
    norm_prov = normalize_name(province)
    norm_city = normalize_name(city)
    norm_dist = normalize_name(district) if district else ''
    matched = None
    match_level = None
    for t in templates:
        if normalize_name(t['province']) == norm_prov and normalize_name(t['city']) == norm_city and normalize_name(t.get('district', '')) == norm_dist and t['report_type'] == report_type:
            matched = t; match_level = "区级模板"; break
    if not matched:
        for t in templates:
            if normalize_name(t['province']) == norm_prov and normalize_name(t['city']) == norm_city and t['report_type'] == report_type:
                matched = t; match_level = "市级模板"; break
    if not matched:
        for t in templates:
            if normalize_name(t['province']) == norm_prov and t['report_type'] == report_type:
                matched = t; match_level = "省级模板"; break
    candidates = [t for t in templates if normalize_name(t['province']) == norm_prov and t['report_type'] == report_type]
    return matched, match_level, candidates

def recommend_template_with_score(province, city, report_type, templates):
    if not templates:
        return None, "无可用模板"
    norm_prov = normalize_name(province)
    norm_city = normalize_name(city)
    candidates = []
    for t in templates:
        score = 0
        match_type = ""
        if normalize_name(t['province']) == norm_prov and normalize_name(t['city']) == norm_city and t['report_type'] == report_type:
            score += 100
            match_type = "精确匹配"
        elif normalize_name(t['city']) == norm_city and t['report_type'] == report_type:
            score += 50
            match_type = "城市级匹配"
        elif normalize_name(t['province']) == norm_prov and t['report_type'] == report_type:
            score += 30
            match_type = "省级匹配"
        score += t.get('use_count', 0) * 0.1
        candidates.append((t, score, match_type))
    if candidates:
        candidates.sort(key=lambda x: x[1], reverse=True)
        best, best_score, match_type = candidates[0]
        if best_score >= 30:
            return best, f"✅ {match_type}（使用{best.get('use_count',0)}次）"
    return None, "💡 无匹配，建议使用通用模板"

def get_custom_template_field_mapping(custom_template):
    if not custom_template:
        return {}
    try:
        return json.loads(custom_template.get('field_mapping', '{}'))
    except:
        return {}

def auto_load_sheet_with_header_detection(file, sheet_name):
    xls = pd.ExcelFile(file)
    df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None)
    header_row = None
    for i, row in df_raw.iterrows():
        row_text = ' '.join([str(v) for v in row.values if pd.notna(v)])
        if ('城市' in row_text and '公司' in row_text) or ('省份' in row_text and '城市' in row_text):
            header_row = i; break
    if header_row is not None:
        df = pd.read_excel(file, sheet_name=sheet_name, skiprows=header_row)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how='all')
        return df, header_row
    else:
        df = pd.read_excel(file, sheet_name=sheet_name, header=0)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how='all')
        return df, 0

def is_valid_chinese_name(text):
    if not text or not isinstance(text, str):
        return False
    return len(re.findall(r'[\u4e00-\u9fa5]', text)) >= 2

def is_valid_city_name(text):
    if not text or not isinstance(text, str):
        return False
    return len(re.findall(r'[\u4e00-\u9fa5]', text)) >= 2

CITY_KEYWORDS = ['城市', '所属城市', '城市名称', '地区', '所属地区', 'city', '地区名', '城市名']
COMPANY_KEYWORDS = ['公司', '分公司', '公司名称', '企业名称', '单位名称', 'company', '单位', '企业']
DISTRICT_KEYWORDS = ['区县', '区', '县', '城区', 'district', '区域']

def detect_columns(df):
    city_col = None
    company_col = None
    district_col = None
    for col in df.columns:
        col_lower = col.lower().strip()
        if any(kw in col_lower for kw in CITY_KEYWORDS):
            city_col = col
        elif any(kw in col_lower for kw in COMPANY_KEYWORDS):
            company_col = col
        elif any(kw in col_lower for kw in DISTRICT_KEYWORDS):
            district_col = col
    return city_col, company_col, district_col

def highlight_error_rows(df, error_rows):
    if not error_rows:
        return df.style
    if isinstance(error_rows, list):
        error_rows = {row: ["异常"] for row in error_rows}
    def row_style(row):
        idx = row.name + 1
        if idx in error_rows:
            return ['background-color: #ffcccc; color: #990000'] * len(row)
        else:
            return [''] * len(row)
    return df.style.apply(row_style, axis=1)

def parse_companies_from_sheet(file, sheet_name, user_id="default"):
    try:
        df = pd.read_excel(file, sheet_name=sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how='all')
        mapping = load_column_mappings(user_id)
        if mapping:
            city_col = mapping.get('city_col')
            company_col = mapping.get('company_col')
            district_col = mapping.get('district_col')
            if city_col and city_col in df.columns and company_col and company_col in df.columns:
                pass
            else:
                city_col, company_col, district_col = detect_columns(df)
        else:
            city_col, company_col, district_col = detect_columns(df)
        if not city_col or not company_col:
            return [], set(), list(df.columns)
        city_province_map = {}
        for r in load_rules():
            key = normalize_name(r['city'])
            city_province_map[key] = r['province']
        companies = []
        unmapped_cities = set()
        for _, row in df.iterrows():
            city = str(row[city_col]) if pd.notna(row[city_col]) else ''
            company = str(row[company_col]) if pd.notna(row[company_col]) else ''
            district = str(row[district_col]) if district_col and pd.notna(row[district_col]) else ''
            if not city or not company:
                continue
            norm_city = normalize_name(city)
            province = city_province_map.get(norm_city, '')
            if not province:
                unmapped_cities.add(city)
            companies.append({
                'company_name': company,
                'province': province,
                'city': city,
                'district': district,
                'tax_id': ''
            })
        return companies, unmapped_cities, list(df.columns)
    except Exception as e:
        st.error(f"解析Sheet失败: {e}")
        return [], set(), []

def validate_data(df, rules):
    if df is None or df.empty:
        return {'total_rows': 0, 'error_rows': 0, 'details': {}, 'summary': {}, 'error_rows_detail': {}}
    errors = {}
    city_col = None
    for col in df.columns:
        if any(kw in col.lower() for kw in CITY_KEYWORDS):
            city_col = col; break
    if city_col is None:
        errors['城市列缺失'] = ['未找到城市列']
        return {'total_rows': len(df), 'error_rows': 0, 'details': errors, 'summary': {'城市列缺失': 1}, 'error_rows_detail': {}}
    city_rule_map = {normalize_name(r['city']): r for r in rules}
    error_rows = {}
    numeric_cols = []
    for col in df.columns:
        col_lower = col.lower()
        if ('基数' in col_lower or '金额' in col_lower or '费用' in col_lower or '比例' in col_lower):
            if not any(w in col_lower for w in ['校验', '状态', '类型', '说明', '备注', '合规', '是否', '判断', '结果']):
                numeric_cols.append(col)
    for idx, row in df.iterrows():
        row_errors = []
        city = str(row[city_col]) if pd.notna(row[city_col]) else ''
        if not city:
            row_errors.append('城市为空')
        else:
            norm_city = normalize_name(city)
            if norm_city not in city_rule_map:
                row_errors.append(f'城市"{city}"未在规则库中（将使用默认值）')
        for col in numeric_cols:
            val = row[col]
            if pd.notna(val):
                try:
                    num = float(str(val).replace(',', ''))
                    if num < 0:
                        row_errors.append(f'列"{col}"值为负数({num})')
                except:
                    row_errors.append(f'列"{col}"值无法转换为数字')
        company_col = None
        for col in df.columns:
            if any(kw in col.lower() for kw in COMPANY_KEYWORDS):
                company_col = col; break
        if company_col is not None:
            val = row[company_col]
            if pd.isna(val) or str(val).strip() == '':
                row_errors.append('公司名称为空')
        if row_errors:
            error_rows[idx+1] = row_errors
    summary = {}
    for row_errors in error_rows.values():
        for err in row_errors:
            if '城市' in err:
                summary['城市问题'] = summary.get('城市问题', 0) + 1
            elif '空' in err:
                summary['空值'] = summary.get('空值', 0) + 1
            elif '数字' in err:
                summary['格式问题'] = summary.get('格式问题', 0) + 1
            else:
                summary['其他'] = summary.get('其他', 0) + 1
    return {
        'total_rows': len(df),
        'error_rows': len(error_rows),
        'details': error_rows,
        'summary': summary,
        'error_rows_detail': error_rows
    }

def get_rule_for_city(city, province=None):
    if not city:
        return None
    rules = load_rules()
    norm_city = normalize_name(city)
    for r in rules:
        if normalize_name(r['city']) == norm_city:
            return r
    if province:
        norm_prov = normalize_name(province)
        for r in rules:
            if normalize_name(r.get('province', '')) == norm_prov:
                fallback = r.copy()
                fallback['source_quote'] = f"省份默认（{province}）"
                return fallback
    new_rule = {
        'id': str(uuid.uuid4())[:8],
        'city': city,
        'province': province or '未知',
        'unit_social': 0.16,
        'personal_social': 0.08,
        'unit_fund': 0.12,
        'personal_fund': 0.12,
        'social_min': 0,
        'social_max': 999999,
        'fund_min': 0,
        'fund_max': 999999,
        'source_quote': '全局默认（自动创建）',
        'is_default': 0,
        'rule_version': 'v1.0',
        'effective_date': datetime.now().strftime('%Y-%m-%d'),
        'source_url': '#',
        'source_title': f'{city}自动创建规则',
        'source_publish_date': datetime.now().strftime('%Y-%m-%d'),
        'collected_at': datetime.now().isoformat(),
        'applicable_region': province or '全国',
        'official_channel': '系统自动创建',
        'notes': f'城市{city}未找到规则，自动创建默认规则',
        'use_count': 0
    }
    rules.append(new_rule)
    save_rules(rules)
    log_audit("系统", "AUTO_CREATE_RULE", "rule", new_rule['id'], f"自动创建规则: {city}")
    return new_rule

def batch_create_missing_rules():
    companies = load_companies()
    if not companies:
        return 0, "请先导入公司数据"
    existing_cities = {normalize_name(r['city']) for r in load_rules()}
    added = 0
    errors = []
    for comp in companies:
        city = comp.get('city', '')
        if not city:
            continue
        norm_city = normalize_name(city)
        if norm_city not in existing_cities:
            province = comp.get('province', '')
            get_rule_for_city(city, province)
            added += 1
            existing_cities.add(norm_city)
    return added, f"已补全 {added} 个城市规则"

def get_data_source_info(df):
    info = {}
    if df is not None and not df.empty:
        for col in df.columns:
            col_lower = str(col).lower()
            if '年份' in col_lower or '年度' in col_lower:
                info['year'] = df[col].iloc[0] if not df[col].empty else '2025'
            if '月份' in col_lower or '月' in col_lower:
                if '统计月份' in col_lower or '月份' in col_lower:
                    info['month'] = df[col].iloc[0] if not df[col].empty else '12'
    return info

def generate_batch_id():
    return f"BATCH_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{str(uuid.uuid4())[:4]}"

def get_verification_progress(batch_id):
    status = load_verification_status(batch_id)
    if not status:
        return {'total': 4, 'completed': 0, 'items': []}
    items = [
        {'name': '官方来源已核对', 'key': 'source_verified', 'done': status.get('source_verified', 0)},
        {'name': '模板版本已核对', 'key': 'template_verified', 'done': status.get('template_verified', 0)},
        {'name': '规则已人工复核', 'key': 'rule_verified', 'done': status.get('rule_verified', 0)},
        {'name': '数据已核对', 'key': 'data_verified', 'done': status.get('data_verified', 0)},
    ]
    completed = sum(1 for item in items if item['done'])
    return {'total': len(items), 'completed': completed, 'items': items}

try:
    import pdfkit
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

def df_to_html_table(df):
    if df is None or df.empty:
        return "<p>无数据</p>"
    html = "<table border='1' style='border-collapse:collapse;font-family:SimSun,serif;'>"
    html += "<thead><tr>"
    for col in df.columns:
        html += f"<th style='background:#4472C4;color:white;padding:8px;'>{col}</th>"
    html += "</tr></thead><tbody>"
    for _, row in df.iterrows():
        html += "<tr>"
        for val in row:
            html += f"<td style='padding:6px;'>{val}</td>"
        html += "</tr>"
    html += "</tbody></table>"
    return html

def create_pdf_from_html(html_content, title="报表"):
    full_html = f"""
    <html>
    <head><meta charset="UTF-8"><title>{title}</title></head>
    <body style='font-family: SimSun, serif; padding: 20px;'>
        <h2 style='text-align:center;'>{title}</h2>
        <p style='text-align:center;color:#666;'>生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <hr/>
        {html_content}
        <hr/>
        <p style='text-align:center;color:#999;font-size:12px;'>由智能报表系统生成</p>
    </body>
    </html>
    """
    return full_html

def show_email_config():
    st.subheader("📧 邮件发送配置")
    config = load_email_config()
    with st.form(key="email_config_form"):
        smtp_server = st.text_input("SMTP服务器", value=config.get('smtp_server', 'smtp.qq.com') if config else 'smtp.qq.com')
        smtp_port = st.number_input("SMTP端口", value=config.get('smtp_port', 465) if config else 465, step=1)
        sender_email = st.text_input("发件邮箱", value=config.get('sender_email', '') if config else '')
        sender_password = st.text_input("邮箱授权码", value=config.get('sender_password', '') if config else '', type="password")
        recipient_email = st.text_input("默认收件人", value=config.get('recipient_email', '') if config else '')
        submitted = st.form_submit_button("保存配置")
        if submitted:
            save_email_config({
                'smtp_server': smtp_server,
                'smtp_port': smtp_port,
                'sender_email': sender_email,
                'sender_password': sender_password,
                'recipient_email': recipient_email
            })
            st.success("邮件配置已保存！")
            log_audit("系统", "SAVE_EMAIL_CONFIG", "config", "", "保存邮件配置")
            st.rerun()

def show_data_versions():
    st.subheader("📊 数据版本历史")
    versions = load_data_versions()
    if versions:
        df = pd.DataFrame(versions)
        st.dataframe(df[['version_number', 'created_at', 'record_count', 'file_name', 'notes']], use_container_width=True)
        selected_version = st.selectbox("选择版本下载", [f"v{v['version_number']} - {v['created_at'][:16]}" for v in versions])
        if selected_version:
            idx = int(selected_version.split(' - ')[0][1:]) - 1
            if idx < len(versions):
                v = versions[idx]
                st.download_button(f"📥 下载 v{v['version_number']}", data=BytesIO(v['file_data']), file_name=v['file_name'])
    else:
        st.info("暂无数据版本记录")

st.set_page_config(page_title="智能报表系统 - 企业版", layout="wide")
st.title("📋 智能报表系统（企业版）")
st.markdown("**工作台 · 数据识别 · 依据库 · 复核导出 · 作业记录 · 年审数据处理 · 审计日志**")

if 'filtered_values' in st.session_state and st.session_state['filtered_values']:
    with st.expander("⚠️ 已过滤的异常数据"):
        for val in st.session_state['filtered_values']:
            st.write(f"- {val}")
    st.session_state['filtered_values'] = []

st.sidebar.title("📌 导航")
page = st.sidebar.radio("选择功能", [
    "📊 工作台",
    "📤 数据导入",
    "📚 依据库管理",
    "⚙️ 规则管理",
    "📄 自定义模板",
    "📋 导出历史与复核",
    "💾 备份与恢复",
    "📋 年审数据处理",
    "🔐 审计日志",
    "📧 邮件发送配置",
    "📊 数据版本对比"
])

if page == "📊 工作台":
    st.subheader("📊 工作台概览")
    col1, col2, col3, col4, col5, col6 = st.columns(6)
    companies = load_companies()
    templates = load_templates()
    rules = load_rules()
    history = load_export_history()
    batches = load_job_batches()
    sources = load_source_registry()
    with col1:
        st.metric("🏢 公司数", len(companies))
    with col2:
        st.metric("📄 模板数", len(templates))
    with col3:
        st.metric("⚙️ 规则数", len(rules))
    with col4:
        pending = len([h for h in history if h.get('review_status') == 'pending'])
        st.metric("📋 待复核", pending)
    with col5:
        st.metric("📚 来源数", len(sources))
    with col6:
        logs = load_audit_log(1)
        st.metric("📋 操作记录", len(load_audit_log(100)))
    st.markdown("---")
    st.subheader("📋 最近作业记录")
    if batches:
        df_batches = pd.DataFrame(batches)
        st.dataframe(df_batches[['batch_name', 'created_at', 'status', 'total_companies', 'total_reports', 'review_status']].head(10), use_container_width=True)
    else:
        st.info("暂无作业记录")
    st.subheader("📊 待处理异常")
    if 'imported_df' in st.session_state and st.session_state['imported_df'] is not None:
        df = st.session_state['imported_df']
        rules = load_rules()
        report = validate_data(df, rules)
        if report['error_rows'] > 0:
            st.warning(f"⚠️ 当前数据有 {report['error_rows']} 行异常")
            with st.expander("查看异常详情"):
                for row, errs in report['details'].items():
                    st.write(f"**第 {row} 行**：{', '.join(errs)}")
        else:
            st.success("✅ 当前数据无异常")
    else:
        st.info("暂无数据，请先导入")

elif page == "📤 数据导入":
    st.subheader("📤 数据导入（支持多文件）")
    uploaded_files = st.file_uploader("选择Excel文件（支持多个 .xlsx）", type=["xlsx"], accept_multiple_files=True)
    if uploaded_files:
        with st.spinner("正在解析Excel..."):
            first_file = uploaded_files[0]
            xls = pd.ExcelFile(first_file)
            all_sheets = xls.sheet_names
            st.session_state['uploaded_files'] = uploaded_files
            st.session_state['all_sheets'] = all_sheets
            default_sheet = None
            for kw in ['年检主数据表', '公司维度明细台账', '主数据', '明细台账']:
                for s in all_sheets:
                    if kw in s:
                        default_sheet = s
                        break
                if default_sheet:
                    break
            if not default_sheet:
                default_sheet = all_sheets[0]
            st.session_state['selected_sheet'] = default_sheet
            companies, unmapped, columns = parse_companies_from_sheet(first_file, default_sheet, user_id="default")
            if not companies and columns:
                st.warning("未能自动识别列，请手动选择")
                city_col = st.selectbox("请选择城市列", [""] + columns)
                company_col = st.selectbox("请选择公司列", [""] + columns)
                district_col = st.selectbox("请选择区县列（可选）", [""] + columns)
                if city_col and company_col:
                    df = pd.read_excel(first_file, sheet_name=default_sheet)
                    city_province_map = {}
                    for r in load_rules():
                        key = normalize_name(r['city'])
                        city_province_map[key] = r['province']
                    companies = []
                    unmapped = set()
                    for _, row in df.iterrows():
                        city = str(row[city_col]) if pd.notna(row[city_col]) else ''
                        company = str(row[company_col]) if pd.notna(row[company_col]) else ''
                        district = str(row[district_col]) if district_col and pd.notna(row[district_col]) else ''
                        if city and company:
                            norm_city = normalize_name(city)
                            province = city_province_map.get(norm_city, '')
                            if not province:
                                unmapped.add(city)
                            companies.append({
                                'company_name': company,
                                'province': province,
                                'city': city,
                                'district': district,
                                'tax_id': ''
                            })
                    save_column_mapping("default", default_sheet, city_col, company_col, district_col)
            if companies:
                save_companies(companies)
                log_audit("系统", "UPLOAD", "companies", "", f"导入 {len(companies)} 家公司")
                st.success(f"✅ 成功提取 {len(companies)} 家公司（Sheet: {default_sheet}）")
                if unmapped:
                    st.warning(f"⚠️ 以下城市未在规则库中找到：{', '.join(unmapped)}，将自动创建默认规则")
                    for city in unmapped:
                        get_rule_for_city(city, '')
                    st.success(f"✅ 已自动创建 {len(unmapped)} 个城市的默认规则")
                df, header_row = auto_load_sheet_with_header_detection(first_file, default_sheet)
                st.session_state['imported_df'] = df
                st.session_state['data_sheet_name'] = default_sheet
                st.session_state['data_header_row'] = header_row
                rules = load_rules()
                st.session_state['validation_report'] = validate_data(df, rules)
                st.session_state['data_recognition'] = {
                    'sheets': all_sheets,
                    'selected_sheet': default_sheet,
                    'header_row': header_row,
                    'columns': list(df.columns) if df is not None else [],
                    'total_rows': len(df) if df is not None else 0,
                    'companies_extracted': len(companies)
                }
                save_data_version(len(companies), first_file.name, first_file.getvalue(), f"导入 {len(companies)} 家公司")
            else:
                st.error("未提取到任何公司数据，请检查Sheet是否正确")
    if 'uploaded_files' in st.session_state and st.session_state['uploaded_files']:
        st.markdown("---")
        st.subheader("📂 选择数据Sheet（切换后自动重新加载公司数据）")
        first_file = st.session_state['uploaded_files'][0]
        xls = pd.ExcelFile(first_file)
        sheets = xls.sheet_names
        current_sheet = st.session_state.get('selected_sheet', sheets[0] if sheets else '')
        selected_sheet = st.selectbox(
            "请选择包含公司数据的Sheet",
            sheets,
            index=sheets.index(current_sheet) if current_sheet in sheets else 0,
            key="sheet_selector_in_import_v2"
        )
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("🔄 重新加载此Sheet数据", key="reload_sheet_data"):
                st.session_state['selected_sheet'] = selected_sheet
                companies, unmapped, columns = parse_companies_from_sheet(first_file, selected_sheet, user_id="default")
                if companies:
                    save_companies(companies)
                    st.success(f"✅ 已重新加载 {len(companies)} 家公司（Sheet: {selected_sheet}）")
                    if unmapped:
                        for city in unmapped:
                            get_rule_for_city(city, '')
                        st.success(f"✅ 已自动创建 {len(unmapped)} 个城市的默认规则")
                    df, header_row = auto_load_sheet_with_header_detection(first_file, selected_sheet)
                    st.session_state['imported_df'] = df
                    st.session_state['data_sheet_name'] = selected_sheet
                    st.session_state['data_header_row'] = header_row
                    rules = load_rules()
                    st.session_state['validation_report'] = validate_data(df, rules)
                    st.session_state['data_recognition']['selected_sheet'] = selected_sheet
                    st.session_state['data_recognition']['total_rows'] = len(df)
                    st.rerun()
                else:
                    st.error("此Sheet未提取到公司数据，请选择其他Sheet")
        with col_btn2:
            if st.button("📊 查看当前公司列表", key="view_companies"):
                companies = load_companies()
                st.dataframe(pd.DataFrame(companies))
                st.caption(f"共 {len(companies)} 家公司")
    st.markdown("---")
    st.subheader("🔍 数据识别结果")
    if 'data_recognition' in st.session_state:
        rec = st.session_state['data_recognition']
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📄 检测到Sheet数", len(rec.get('sheets', [])))
        with col2:
            st.metric("📊 当前Sheet", rec.get('selected_sheet', '无'))
        with col3:
            st.metric("📋 表头行", rec.get('header_row', 0) + 1 if rec.get('header_row') is not None else 1)
        with col4:
            st.metric("🏢 提取公司数", rec.get('companies_extracted', 0))
        if rec.get('columns'):
            st.write("**表头字段**")
            st.dataframe(pd.DataFrame({'字段名': rec['columns']}), use_container_width=True)
        if rec.get('sheets'):
            st.write("**所有Sheet**")
            st.write(", ".join(rec['sheets']))
        if st.button("🗑️ 清空识别结果", key="clear_recognition"):
            st.session_state['data_recognition'] = {}
            st.rerun()
    else:
        st.info("请上传文件以查看数据识别结果")
    if 'imported_df' in st.session_state and st.session_state['imported_df'] is not None:
        st.subheader("📊 数据预览")
        df = st.session_state['imported_df']
        if 'validation_report' in st.session_state:
            report = st.session_state['validation_report']
            error_rows = report.get('details', {})
            styled_df = highlight_error_rows(df.head(10), error_rows)
            st.dataframe(styled_df, use_container_width=True)
            st.caption("🔴 红色行表示存在数据异常")
        else:
            st.dataframe(df.head(10), use_container_width=True)
        st.caption(f"当前Sheet: {st.session_state.get('data_sheet_name', '未知')}，共 {len(df)} 行")
        if st.button("📤 导出清洗后数据（Excel）", key="export_cleaned"):
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='清洗后数据', index=False)
            output.seek(0)
            st.download_button("📥 下载清洗后数据", data=output, file_name=f"清洗后数据_{datetime.now().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            log_audit("系统", "EXPORT_CLEANED", "data", "", "导出清洗后数据")
        if 'validation_report' in st.session_state:
            report = st.session_state['validation_report']
            st.subheader("📊 数据质量报告")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("总行数", report['total_rows'])
            with col2:
                st.metric("异常行数", report['error_rows'])
            with col3:
                st.metric("正常行数", report['total_rows'] - report['error_rows'])
            if report['error_rows'] > 0:
                st.warning(f"⚠️ 发现 {report['error_rows']} 行数据存在问题")
                with st.expander("查看异常详情"):
                    for row, errs in report['details'].items():
                        st.write(f"**第 {row} 行**：{', '.join(errs)}")
            else:
                st.success("✅ 所有数据校验通过！")

elif page == "📚 依据库管理":
    st.subheader("📚 依据库管理")
    tab1, tab2, tab3, tab4 = st.tabs(["📄 模板", "⚙️ 规则", "🏢 公司", "📚 来源"])
    with tab1:
        templates = load_templates()
        if templates:
            df = pd.DataFrame(templates)
            cols = ['template_name', 'report_type', 'province', 'city', 'template_version', 'source_authority', 'use_count']
            st.dataframe(df[cols], use_container_width=True)
        else:
            st.info("暂无模板")
    with tab2:
        rules = load_rules()
        if rules:
            df = pd.DataFrame(rules)
            cols = ['city', 'province', 'unit_social', 'personal_social', 'unit_fund', 'personal_fund', 'source_quote', 'rule_version', 'use_count']
            st.dataframe(df[cols], use_container_width=True)
        else:
            st.info("暂无规则")
    with tab3:
        companies = load_companies()
        if companies:
            st.dataframe(pd.DataFrame(companies))
            st.caption(f"共 {len(companies)} 家公司")
            st.markdown("**批量操作**")
            selected_for_delete = st.multiselect("选择要删除的公司", [c['company_name'] for c in companies])
            if st.button("🗑️ 删除选中公司", key="batch_delete_companies"):
                if selected_for_delete:
                    remaining = [c for c in companies if c['company_name'] not in selected_for_delete]
                    save_companies(remaining)
                    log_audit("系统", "BATCH_DELETE", "companies", "", f"删除 {len(selected_for_delete)} 家公司")
                    st.success(f"已删除 {len(selected_for_delete)} 家公司")
                    st.rerun()
        else:
            st.info("暂无公司数据")
    with tab4:
        st.markdown("**📚 官方来源管理**")
        sources = load_source_registry()
        with st.expander("➕ 新增官方来源", expanded=False):
            with st.form(key="add_source_form"):
                col1, col2 = st.columns(2)
                with col1:
                    new_id = st.text_input("来源ID", value=f"src_{str(uuid.uuid4())[:8]}")
                    new_authority_type = st.selectbox("机构类型", ['social_security', 'tax', 'fund', 'other'])
                    new_province = st.text_input("省份")
                    new_city = st.text_input("城市")
                    new_district = st.text_input("区县")
                    new_authority_name = st.text_input("机构名称")
                    new_official_site_name = st.text_input("官网名称")
                with col2:
                    new_source_url = st.text_input("来源URL")
                    new_source_level = st.selectbox("来源级别", ['city', 'province', 'national'])
                    new_source_section = st.text_input("来源板块")
                    new_is_official = st.checkbox("官方来源", value=True)
                    new_crawl_allowed = st.checkbox("允许爬取", value=True)
                    new_document_name = st.text_input("文档名称")
                    new_document_version = st.text_input("文档版本")
                    new_publish_year = st.text_input("发布年份")
                    new_notes = st.text_area("备注")
                submitted = st.form_submit_button("保存来源")
                if submitted:
                    new_source = {
                        'id': new_id,
                        'authority_type': new_authority_type,
                        'province': new_province,
                        'city': new_city,
                        'district': new_district,
                        'authority_name': new_authority_name,
                        'official_site_name': new_official_site_name,
                        'source_url': new_source_url,
                        'source_level': new_source_level,
                        'source_section': new_source_section,
                        'is_official': 1 if new_is_official else 0,
                        'crawl_allowed': 1 if new_crawl_allowed else 0,
                        'last_checked': datetime.now().strftime('%Y-%m-%d'),
                        'status': 'active',
                        'notes': new_notes,
                        'document_name': new_document_name,
                        'document_version': new_document_version,
                        'publish_year': new_publish_year
                    }
                    sources.append(new_source)
                    save_source_registry(sources)
                    log_audit("系统", "ADD", "source", new_id, f"新增来源: {new_authority_name}")
                    st.success("来源已添加！")
                    st.rerun()
        if sources:
            df_sources = pd.DataFrame(sources)
            display_cols = ['id', 'authority_type', 'province', 'city', 'authority_name', 'source_url', 'document_name', 'document_version']
            st.dataframe(df_sources[display_cols], use_container_width=True)
            st.markdown("**操作**")
            col_ops1, col_ops2 = st.columns(2)
            with col_ops1:
                source_to_delete = st.selectbox("选择要删除的来源ID", [s['id'] for s in sources] if sources else [])
                if st.button("🗑️ 删除所选来源"):
                    if source_to_delete:
                        delete_source_by_id(source_to_delete)
                        log_audit("系统", "DELETE", "source", source_to_delete, f"删除来源")
                        st.success(f"已删除来源 {source_to_delete}")
                        st.rerun()
            with col_ops2:
                source_to_edit = st.selectbox("选择要编辑的来源ID", [s['id'] for s in sources] if sources else [])
                if source_to_edit:
                    with st.expander(f"✏️ 编辑来源 {source_to_edit}", expanded=False):
                        source = next(s for s in sources if s['id'] == source_to_edit)
                        with st.form(key="edit_source_form"):
                            col1, col2 = st.columns(2)
                            with col1:
                                edit_authority_type = st.selectbox("机构类型", ['social_security', 'tax', 'fund', 'other'], index=['social_security','tax','fund','other'].index(source.get('authority_type','social_security')))
                                edit_province = st.text_input("省份", value=source.get('province',''))
                                edit_city = st.text_input("城市", value=source.get('city',''))
                                edit_district = st.text_input("区县", value=source.get('district',''))
                                edit_authority_name = st.text_input("机构名称", value=source.get('authority_name',''))
                                edit_official_site_name = st.text_input("官网名称", value=source.get('official_site_name',''))
                            with col2:
                                edit_source_url = st.text_input("来源URL", value=source.get('source_url',''))
                                edit_source_level = st.selectbox("来源级别", ['city', 'province', 'national'], index=['city','province','national'].index(source.get('source_level','city')))
                                edit_source_section = st.text_input("来源板块", value=source.get('source_section',''))
                                edit_is_official = st.checkbox("官方来源", value=bool(source.get('is_official',1)))
                                edit_crawl_allowed = st.checkbox("允许爬取", value=bool(source.get('crawl_allowed',1)))
                                edit_document_name = st.text_input("文档名称", value=source.get('document_name',''))
                                edit_document_version = st.text_input("文档版本", value=source.get('document_version',''))
                                edit_publish_year = st.text_input("发布年份", value=source.get('publish_year',''))
                                edit_notes = st.text_area("备注", value=source.get('notes',''))
                            submitted_edit = st.form_submit_button("更新来源")
                            if submitted_edit:
                                updated_sources = []
                                for s in sources:
                                    if s['id'] == source_to_edit:
                                        s.update({
                                            'authority_type': edit_authority_type,
                                            'province': edit_province,
                                            'city': edit_city,
                                            'district': edit_district,
                                            'authority_name': edit_authority_name,
                                            'official_site_name': edit_official_site_name,
                                            'source_url': edit_source_url,
                                            'source_level': edit_source_level,
                                            'source_section': edit_source_section,
                                            'is_official': 1 if edit_is_official else 0,
                                            'crawl_allowed': 1 if edit_crawl_allowed else 0,
                                            'last_checked': datetime.now().strftime('%Y-%m-%d'),
                                            'document_name': edit_document_name,
                                            'document_version': edit_document_version,
                                            'publish_year': edit_publish_year,
                                            'notes': edit_notes
                                        })
                                    updated_sources.append(s)
                                save_source_registry(updated_sources)
                                log_audit("系统", "EDIT", "source", source_to_edit, f"编辑来源: {edit_authority_name}")
                                st.success(f"来源 {source_to_edit} 已更新！")
                                st.rerun()
        else:
            st.info("暂无来源，请添加或加载样本依据")
        if st.button("加载样本依据"):
            sample_sources = [
                {'id': 'src_gz_social_demo', 'authority_type': 'social_security', 'province': '广东', 'city': '广州', 'district': '', 'authority_name': '广州市人力资源和社会保障局', 'official_site_name': '广州市人力资源和社会保障局官网', 'source_url': 'https://rsj.gz.gov.cn/', 'source_level': 'city', 'source_section': '网上办事', 'document_name': '广州市社保年审公告', 'document_version': '2024', 'publish_year': '2024', 'is_official': 1, 'crawl_allowed': 1, 'last_checked': datetime.now().strftime('%Y-%m-%d'), 'status': 'active', 'notes': '示例，请核实'},
                {'id': 'src_sh_social_sample', 'authority_type': 'social_security', 'province': '上海', 'city': '上海', 'district': '', 'authority_name': '上海市人力资源和社会保障局', 'official_site_name': '上海社保官方系统', 'source_url': 'https://rsj.sh.gov.cn/', 'source_level': 'city', 'source_section': '网上办事', 'document_name': '上海市社保基数调整通知', 'document_version': '2024', 'publish_year': '2024', 'is_official': 1, 'crawl_allowed': 1, 'last_checked': datetime.now().strftime('%Y-%m-%d'), 'status': 'active', 'notes': '示例，请核实'},
                {'id': 'src_hrb_social_sample', 'authority_type': 'social_security', 'province': '黑龙江', 'city': '哈尔滨', 'district': '', 'authority_name': '哈尔滨市人力资源和社会保障局', 'official_site_name': '哈尔滨社保官方系统', 'source_url': '', 'source_level': 'city', 'source_section': '网上办事', 'document_name': '哈尔滨年审公告', 'document_version': '2024', 'publish_year': '2024', 'is_official': 1, 'crawl_allowed': 1, 'last_checked': datetime.now().strftime('%Y-%m-%d'), 'status': 'active', 'notes': '示例，请核实'}
            ]
            save_source_registry(sample_sources)
            log_audit("系统", "LOAD_SAMPLE", "source", "", "加载样本依据")
            st.success("已加载样本依据！")
            st.rerun()

elif page == "⚙️ 规则管理":
    st.subheader("⚙️ 规则管理（社保/公积金）")
    st.markdown("**⚠️ 重要：所有规则必须来自官方渠道，并填写来源信息。**")
    st.markdown("官方渠道优先级：国家税务总局、各省税务局、人社部、各城市公积金中心等官网。")
    st.markdown("**💡 提示：如果城市未找到规则，系统会自动创建默认规则（16%/8%）并入库。**")
    rules = load_rules()
    st.write(f"**当前规则数量：{len(rules)} 个城市**")
    if rules:
        df_rules = pd.DataFrame(rules)
        st.dataframe(df_rules[['city', 'province', 'unit_social', 'personal_social', 'unit_fund', 'personal_fund', 'social_min', 'social_max', 'source_quote', 'rule_version', 'use_count']], use_container_width=True)
        
        col_btn1, col_btn2, col_btn3 = st.columns(3)
        with col_btn1:
            if st.button("🔧 从公司数据批量补全规则"):
                added, msg = batch_create_missing_rules()
                st.success(msg)
                st.rerun()
        with col_btn2:
            if st.button("📋 导出规则清单"):
                st.dataframe(df_rules[['city', 'province', 'unit_social', 'personal_social', 'unit_fund', 'personal_fund']], use_container_width=True)
        with col_btn3:
            with st.expander("📋 跨城市规则复制", expanded=False):
                if len(rules) >= 2:
                    source_city = st.selectbox("源城市", [r['city'] for r in rules])
                    target_city = st.text_input("目标城市")
                    if st.button("复制规则"):
                        source_rule = next((r for r in rules if r['city'] == source_city), None)
                        if source_rule and target_city:
                            new_rule = source_rule.copy()
                            new_rule['id'] = str(uuid.uuid4())[:8]
                            new_rule['city'] = target_city
                            new_rule['source_quote'] = f"复制自{source_city}"
                            new_rule['notes'] = f"从{source_city}复制"
                            rules.append(new_rule)
                            save_rules(rules)
                            log_audit("系统", "COPY_RULE", "rule", new_rule['id'], f"从{source_city}复制到{target_city}")
                            st.success(f"已复制 {source_city} 规则到 {target_city}")
                            st.rerun()
                else:
                    st.info("至少需要2个城市规则才能复制")
        
        with st.expander("📜 规则变更历史", expanded=False):
            hist = load_rule_history()
            if hist:
                df_hist = pd.DataFrame(hist)
                st.dataframe(df_hist[['rule_id', 'action_type', 'changed_by', 'changed_at', 'notes']], use_container_width=True)
                hist_id = st.selectbox("选择历史版本回滚", [h['id'] for h in hist] if hist else [])
                if st.button("↩️ 回滚到此版本"):
                    if rollback_rule(hist_id):
                        st.success("回滚成功！")
                        st.rerun()
                    else:
                        st.error("回滚失败")
            else:
                st.info("暂无规则变更历史")
        
        with st.expander("➕ 新增城市规则（需填写官方来源）", expanded=False):
            with st.form(key="add_rule_form"):
                col1, col2 = st.columns(2)
                with col1:
                    new_city = st.text_input("城市名称 *")
                    new_province = st.text_input("所属省份 *")
                    new_unit_social = st.number_input("单位社保比例", value=0.16, step=0.001, format="%.3f")
                    new_personal_social = st.number_input("个人社保比例", value=0.08, step=0.001, format="%.3f")
                    new_unit_fund = st.number_input("单位公积金比例", value=0.12, step=0.001, format="%.3f")
                    new_personal_fund = st.number_input("个人公积金比例", value=0.12, step=0.001, format="%.3f")
                    new_social_min = st.number_input("社保基数下限", value=0.0, step=100.0)
                    new_social_max = st.number_input("社保基数上限", value=999999.0, step=100.0)
                    new_fund_min = st.number_input("公积金基数下限", value=0.0, step=100.0)
                    new_fund_max = st.number_input("公积金基数上限", value=999999.0, step=100.0)
                    new_source_quote = st.text_input("来源文号（可选）")
                    new_rule_version = st.text_input("规则版本", value="v1.0")
                    new_effective_date = st.text_input("生效日期", value=datetime.now().strftime('%Y-%m-%d'))
                with col2:
                    st.markdown("**来源信息（必须填写）**")
                    new_source_url = st.text_input("来源链接 *")
                    new_source_title = st.text_input("来源标题 *")
                    new_source_publish_date = st.text_input("来源发布时间 *", value=datetime.now().strftime('%Y-%m-%d'))
                    new_applicable_region = st.text_input("适用地区 *", help="如'全国'或'广东'")
                    new_official_channel = st.selectbox("官方渠道 *", ['国家税务总局官网', '国家税务总局政策法规库', '12366办税指南', '省级税务局官网', '市级/区县级税务局官网', '全国人社政务服务平台', '地方人社局官网', '地方社保中心', '住房公积金管理中心官网', '其他'])
                    new_notes = st.text_area("备注")
                submitted = st.form_submit_button("添加规则")
                if submitted:
                    if not new_city or not new_province or not new_source_url or not new_source_title:
                        st.error("城市、省份、来源链接、来源标题为必填项")
                    else:
                        new_rule = {
                            'id': str(uuid.uuid4())[:8],
                            'city': new_city,
                            'province': new_province,
                            'unit_social': new_unit_social,
                            'personal_social': new_personal_social,
                            'unit_fund': new_unit_fund,
                            'personal_fund': new_personal_fund,
                            'social_min': new_social_min,
                            'social_max': new_social_max,
                            'fund_min': new_fund_min,
                            'fund_max': new_fund_max,
                            'source_quote': new_source_quote,
                            'is_default': 0,
                            'rule_version': new_rule_version,
                            'effective_date': new_effective_date,
                            'source_url': new_source_url,
                            'source_title': new_source_title,
                            'source_publish_date': new_source_publish_date,
                            'collected_at': datetime.now().isoformat(),
                            'applicable_region': new_applicable_region,
                            'official_channel': new_official_channel,
                            'notes': new_notes,
                            'use_count': 0
                        }
                        rules.append(new_rule)
                        save_rules(rules)
                        log_rule_change(new_rule['id'], "CREATE", {}, new_rule, "系统", "新增规则")
                        log_audit("系统", "ADD", "rule", new_rule['id'], f"新增规则: {new_city}")
                        st.success(f"已添加 {new_city} 的规则！")
                        st.rerun()
        cities = sorted(set(r['city'] for r in rules))
        selected_city = st.selectbox("选择城市进行编辑", [""] + cities)
        if selected_city:
            rule = next((r for r in rules if r['city'] == selected_city), None)
            if rule:
                with st.form(key=f"edit_rule_{selected_city}"):
                    st.write(f"编辑 **{selected_city}** 的规则")
                    col1, col2 = st.columns(2)
                    with col1:
                        new_unit_social = st.number_input("单位社保比例", value=float(rule['unit_social']), step=0.001, format="%.3f")
                        new_personal_social = st.number_input("个人社保比例", value=float(rule['personal_social']), step=0.001, format="%.3f")
                        new_unit_fund = st.number_input("单位公积金比例", value=float(rule['unit_fund']), step=0.001, format="%.3f")
                        new_personal_fund = st.number_input("个人公积金比例", value=float(rule['personal_fund']), step=0.001, format="%.3f")
                        new_social_min = st.number_input("社保基数下限", value=float(rule.get('social_min', 0)), step=100.0)
                        new_social_max = st.number_input("社保基数上限", value=float(rule.get('social_max', 999999)), step=100.0)
                        new_fund_min = st.number_input("公积金基数下限", value=float(rule.get('fund_min', 0)), step=100.0)
                        new_fund_max = st.number_input("公积金基数上限", value=float(rule.get('fund_max', 999999)), step=100.0)
                        new_source_quote = st.text_input("来源文号", value=rule.get('source_quote', ''))
                        new_rule_version = st.text_input("规则版本", value=rule.get('rule_version', 'v1.0'))
                        new_effective_date = st.text_input("生效日期", value=rule.get('effective_date', datetime.now().strftime('%Y-%m-%d')))
                    with col2:
                        new_source_url = st.text_input("来源链接", value=rule.get('source_url', ''))
                        new_source_title = st.text_input("来源标题", value=rule.get('source_title', ''))
                        new_source_publish_date = st.text_input("来源发布时间", value=rule.get('source_publish_date', datetime.now().strftime('%Y-%m-%d')))
                        new_applicable_region = st.text_input("适用地区", value=rule.get('applicable_region', ''))
                        new_official_channel = st.selectbox("官方渠道", ['国家税务总局官网', '国家税务总局政策法规库', '12366办税指南', '省级税务局官网', '市级/区县级税务局官网', '全国人社政务服务平台', '地方人社局官网', '地方社保中心', '住房公积金管理中心官网', '其他'], index=['国家税务总局官网','国家税务总局政策法规库','12366办税指南', '省级税务局官网','市级/区县级税务局官网','全国人社政务服务平台', '地方人社局官网','地方社保中心','住房公积金管理中心官网','其他'].index(rule.get('official_channel','其他')))
                        new_notes = st.text_area("备注", value=rule.get('notes', ''))
                    submitted = st.form_submit_button("保存修改")
                    if submitted:
                        old_vals = rule.copy()
                        updated_rules = []
                        for r in rules:
                            if r['id'] == rule['id']:
                                r.update({
                                    'unit_social': new_unit_social,
                                    'personal_social': new_personal_social,
                                    'unit_fund': new_unit_fund,
                                    'personal_fund': new_personal_fund,
                                    'social_min': new_social_min,
                                    'social_max': new_social_max,
                                    'fund_min': new_fund_min,
                                    'fund_max': new_fund_max,
                                    'source_quote': new_source_quote,
                                    'rule_version': new_rule_version,
                                    'effective_date': new_effective_date,
                                    'source_url': new_source_url,
                                    'source_title': new_source_title,
                                    'source_publish_date': new_source_publish_date,
                                    'applicable_region': new_applicable_region,
                                    'official_channel': new_official_channel,
                                    'notes': new_notes,
                                    'collected_at': datetime.now().isoformat()
                                })
                            updated_rules.append(r)
                        save_rules(updated_rules)
                        log_rule_change(rule['id'], "UPDATE", old_vals, r, "系统", f"编辑规则: {selected_city}")
                        log_audit("系统", "EDIT", "rule", rule['id'], f"编辑规则: {selected_city}")
                        st.success("规则已更新！")
                        st.rerun()
        if st.button("🔄 重置所有规则为系统默认值（会覆盖所有自定义规则）"):
            if st.checkbox("确认重置？此操作将覆盖所有自定义规则"):
                init_rules_from_default()
                log_audit("系统", "RESET", "rule", "", "重置所有规则")
                st.success("已重置为默认规则！")
                st.rerun()
    else:
        st.info("暂无规则，请添加")

elif page == "📄 自定义模板":
    st.subheader("📄 自定义模板管理")
    custom_templates = load_custom_templates()
    if custom_templates:
        st.write("**已保存的自定义模板**")
        for ct in custom_templates:
            col1, col2, col3 = st.columns([3, 2, 1])
            with col1:
                st.write(f"📄 {ct['name']}")
            with col2:
                st.write(f"字段数：{len(json.loads(ct.get('field_mapping', '{}')))}")
            with col3:
                if st.button("删除", key=f"del_{ct['id']}"):
                    delete_custom_template(ct['id'])
                    log_audit("系统", "DELETE", "custom_template", ct['id'], f"删除自定义模板: {ct['name']}")
                    st.success(f"已删除 {ct['name']}")
                    st.rerun()
    st.markdown("---")
    st.markdown("**上传新模板**")
    uploaded_template = st.file_uploader("选择模板文件（.xlsx）", type=["xlsx"], key="custom_template_upload")
    if uploaded_template:
        try:
            template_name = st.text_input("模板名称", value=uploaded_template.name.replace('.xlsx', ''))
            sheet_name = st.text_input("Sheet名称（留空使用第一个Sheet）", value="")
            field_mapping_source = st.text_input("字段映射来源说明（可选）")
            wb = load_workbook(BytesIO(uploaded_template.getvalue()))
            ws = wb.active if not sheet_name else wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            st.subheader("📄 模板预览")
            preview_df = pd.DataFrame([cell.value for cell in ws[1] if cell.value])
            if not preview_df.empty:
                st.dataframe(preview_df.head(10), use_container_width=True)
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            st.write(f"**检测到表头字段**：{', '.join(headers) if headers else '未检测到表头'}")
            st.markdown("**字段映射（将数据字段映射到模板列）**")
            mapping = {}
            if headers:
                field_options = [''] + ['纳税人识别号', '公司名称', '销售额', '进项税额', '应纳税额', '单位名称', '社保登记号', '基数', '单位金额', '个人金额', '单位比例', '个人比例', '公积金账号', '收入额', '专项扣除', '营业收入', '营业成本', '应纳税所得额', '全年收入', '全年成本', '已预缴税额', '应补退税额', '申报金额']
                for h in headers:
                    col = st.selectbox(f"字段 '{h}' 映射到", field_options, key=f"map_{h}_{uploaded_template.name}")
                    if col:
                        mapping[h] = col
            if st.button("保存自定义模板"):
                if not template_name:
                    st.error("请填写模板名称")
                else:
                    template_data = {'id': str(uuid.uuid4())[:8], 'name': template_name, 'file_data': uploaded_template.getvalue(), 'field_mapping': mapping, 'sheet_name': sheet_name, 'created_at': datetime.now().isoformat()}
                    save_custom_template(template_data)
                    log_audit("系统", "ADD", "custom_template", template_data['id'], f"新增自定义模板: {template_name}")
                    st.success(f"模板 '{template_name}' 已保存！")
                    st.rerun()
        except Exception as e:
            st.error(f"处理模板失败：{e}")
    with st.expander("📜 模板变更历史", expanded=False):
        hist = load_template_history()
        if hist:
            df_hist = pd.DataFrame(hist)
            st.dataframe(df_hist[['template_id', 'action_type', 'changed_by', 'changed_at', 'notes']], use_container_width=True)
            hist_id = st.selectbox("选择历史版本回滚", [h['id'] for h in hist] if hist else [])
            if st.button("↩️ 回滚到此版本"):
                if rollback_template(hist_id):
                    st.success("回滚成功！")
                    st.rerun()
                else:
                    st.error("回滚失败")
        else:
            st.info("暂无模板变更历史")

elif page == "📋 导出历史与复核":
    st.subheader("📋 导出历史与复核")
    tab1, tab2, tab3 = st.tabs(["📋 导出历史", "✅ 复核处理", "📊 核验状态"])
    with tab1:
        history = load_export_history()
        if history:
            df_hist = pd.DataFrame(history)
            display_cols = ['company_name', 'city', 'report_type', 'period_type', 'data_source', 'generated_at', 'review_status', 'batch_id']
            st.dataframe(df_hist[display_cols], use_container_width=True)
            batches = load_job_batches()
            if batches:
                st.subheader("📦 批次作业")
                st.dataframe(pd.DataFrame(batches), use_container_width=True)
        else:
            st.info("暂无导出记录")
    with tab2:
        history = load_export_history()
        pending = [h for h in history if h.get('review_status') == 'pending']
        if pending:
            st.write(f"共 {len(pending)} 份待复核报表")
            for h in pending:
                col1, col2, col3 = st.columns([3, 2, 1])
                with col1:
                    st.write(f"📄 {h['company_name']} - {h['city']} ({h['report_type']})")
                    st.caption(f"生成时间：{h['generated_at'][:16]}")
                with col2:
                    st.write(f"来源：{h.get('data_source', '未知')}")
                with col3:
                    if st.button("✅ 通过", key=f"approve_{h['id']}"):
                        update_export_status(h['id'], 'approved', '复核员')
                        log_audit("复核员", "APPROVE", "export", h['id'], f"通过复核: {h['company_name']}")
                        st.success("已通过复核")
                        st.rerun()
                    if st.button("❌ 驳回", key=f"reject_{h['id']}"):
                        update_export_status(h['id'], 'rejected', '复核员')
                        log_audit("复核员", "REJECT", "export", h['id'], f"驳回复核: {h['company_name']}")
                        st.warning("已驳回")
                        st.rerun()
        else:
            st.success("✅ 暂无待复核报表")
    with tab3:
        st.markdown("**📊 核验状态**")
        batches = load_job_batches()
        if batches:
            for batch in batches[:10]:
                batch_id = batch['id']
                progress = get_verification_progress(batch_id)
                status = load_verification_status(batch_id)
                st.markdown(f"**批次：{batch.get('batch_name', batch_id)}**")
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.progress(progress['completed'] / progress['total'] if progress['total'] > 0 else 0)
                    st.caption(f"{progress['completed']}/{progress['total']} 项已完成")
                    for item in progress['items']:
                        icon = "✅" if item['done'] else "❌"
                        st.write(f"{icon} {item['name']}")
                with col2:
                    st.write(f"导出类型：{status.get('export_type', '验证版') if status else '验证版'}")
                    if status and status.get('reviewer_name'):
                        st.write(f"复核人：{status['reviewer_name']}")
                    st.write(f"状态：{'✅ 可导出正式版' if progress['completed'] == progress['total'] else '⏳ 待完成核验'}")
                    if progress['completed'] == progress['total']:
                        if st.button(f"📤 导出正式版", key=f"formal_{batch_id}"):
                            st.success(f"正式版导出成功！批次：{batch.get('batch_name', batch_id)}")
                    else:
                        if st.button(f"📤 跳过核验导出验证版", key=f"verify_{batch_id}"):
                            if status:
                                update_verification_status(batch_id, 'export_type', '验证版')
                            st.success(f"验证版导出成功！批次：{batch.get('batch_name', batch_id)}")
                    if status:
                        col_a, col_b = st.columns(2)
                        with col_a:
                            if not status.get('source_verified', 0):
                                if st.button(f"✅ 来源已核对", key=f"src_{batch_id}"):
                                    update_verification_status(batch_id, 'source_verified', 1)
                                    log_audit("系统", "VERIFY", "batch", batch_id, "标记来源已核对")
                                    st.rerun()
                            if not status.get('template_verified', 0):
                                if st.button(f"✅ 模板已核对", key=f"tpl_{batch_id}"):
                                    update_verification_status(batch_id, 'template_verified', 1)
                                    log_audit("系统", "VERIFY", "batch", batch_id, "标记模板已核对")
                                    st.rerun()
                        with col_b:
                            if not status.get('rule_verified', 0):
                                if st.button(f"✅ 规则已复核", key=f"rule_{batch_id}"):
                                    update_verification_status(batch_id, 'rule_verified', 1)
                                    log_audit("系统", "VERIFY", "batch", batch_id, "标记规则已复核")
                                    st.rerun()
                            if not status.get('data_verified', 0):
                                if st.button(f"✅ 数据已核对", key=f"data_{batch_id}"):
                                    update_verification_status(batch_id, 'data_verified', 1)
                                    log_audit("系统", "VERIFY", "batch", batch_id, "标记数据已核对")
                                    st.rerun()
                    with st.form(key=f"reviewer_form_{batch_id}"):
                        reviewer_name = st.text_input("复核人姓名", value=status.get('reviewer_name', '') if status else '')
                        if st.form_submit_button("保存复核人"):
                            if status:
                                update_verification_status(batch_id, 'reviewer_name', reviewer_name)
                                log_audit("系统", "SET_REVIEWER", "batch", batch_id, f"设置复核人: {reviewer_name}")
                                st.success("复核人已保存")
                                st.rerun()
                            else:
                                save_verification_status({'batch_id': batch_id, 'source_verified': 0, 'template_verified': 0, 'rule_verified': 0, 'data_verified': 0, 'reviewer_name': reviewer_name, 'verified_at': datetime.now().isoformat(), 'export_type': '验证版'})
                                st.success("核验状态已创建")
                                st.rerun()
                st.markdown("---")
        else:
            st.info("暂无批次作业，请先生成报表")

elif page == "💾 备份与恢复":
    st.subheader("💾 备份与恢复")
    st.markdown("**自动备份**：每次操作后自动备份，保留最近30份")
    backups = get_backup_list()
    if backups:
        st.write(f"共 {len(backups)} 份备份")
        st.markdown("**恢复到指定时间点**")
        backup_dates = [f.split('_')[2] + '_' + f.split('_')[3] for f in backups]
        selected_date = st.selectbox("选择备份时间点", backup_dates)
        if st.button("🔄 恢复到选定时间点"):
            selected_backup = backups[backup_dates.index(selected_date)]
            if restore_backup(selected_backup):
                log_audit("系统", "RESTORE", "backup", selected_backup, f"恢复到 {selected_date}")
                st.success("恢复成功！请刷新页面")
                st.rerun()
            else:
                st.error("恢复失败")
        st.markdown("---")
        st.write("**全部备份列表**")
        selected_backup = st.selectbox("选择备份恢复", backups)
        if st.button("恢复所选备份"):
            if restore_backup(selected_backup):
                log_audit("系统", "RESTORE", "backup", selected_backup, "恢复备份")
                st.success("恢复成功！请刷新页面")
                st.rerun()
            else:
                st.error("恢复失败")
    else:
        st.info("暂无备份")
    if st.button("手动备份当前数据"):
        path = backup_database()
        if path:
            log_audit("系统", "BACKUP", "database", "", f"手动备份: {os.path.basename(path)}")
            st.success(f"备份成功：{os.path.basename(path)}")

elif page == "📋 年审数据处理":
    st.subheader("📋 年审数据处理与交付")
    st.markdown("**处理流程：系统账单筛选 → 上海数据合并 → 基数核算 → 实缴核算 → 归档交付**")
    st.markdown("### 步骤1：上传数据源")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("**📄 to客户文件**")
        to_client_file = st.file_uploader("上传to客户最终交付文件", type=["xlsx"], key="to_client")
        if to_client_file:
            st.success(f"已上传: {to_client_file.name}")
            st.session_state['to_client_file'] = to_client_file
    with col2:
        st.markdown("**📊 内部系统账单**")
        internal_bill_file = st.file_uploader("上传内部系统账单", type=["xlsx"], key="internal_bill")
        if internal_bill_file:
            st.success(f"已上传: {internal_bill_file.name}")
            st.session_state['internal_bill_file'] = internal_bill_file
    with col3:
        st.markdown("**📚 上海社保数据**")
        shanghai_files = st.file_uploader("上传上海社保数据（可多选）", type=["xlsx"], accept_multiple_files=True, key="shanghai")
        if shanghai_files:
            st.success(f"已上传 {len(shanghai_files)} 份文件")
            st.session_state['shanghai_files'] = shanghai_files
    st.markdown("---")
    if st.button("🔄 执行数据处理", key="process_annual_data"):
        if not st.session_state.get('internal_bill_file'):
            st.error("请先上传内部系统账单")
            st.stop()
        with st.spinner("正在处理数据..."):
            processed_files = {}
            results = {}
            try:
                df_bill = pd.read_excel(st.session_state['internal_bill_file'])
                st.write(f"原始账单行数: {len(df_bill)}")
                city_col = None
                for col in df_bill.columns:
                    if any(kw in col.lower() for kw in ['城市', '地区', '所属地']):
                        city_col = col; break
                if city_col:
                    df_gz = df_bill[df_bill[city_col].astype(str).str.contains('广州', na=False)]
                else:
                    df_gz = df_bill
                    st.warning("未识别到城市列，将使用全部数据")
                st.write(f"筛选后广州员工: {len(df_gz)} 行")
                results['gz_bill'] = df_gz
                gz_buffer = BytesIO()
                with pd.ExcelWriter(gz_buffer, engine='openpyxl') as writer:
                    df_gz.to_excel(writer, sheet_name='广州员工明细', index=False)
                gz_buffer.seek(0)
                processed_files['广州内部账单.xlsx'] = gz_buffer.getvalue()
                log_audit("系统", "ANNUAL", "gz_bill", "", f"筛选广州员工 {len(df_gz)} 行")
                st.success(f"✅ 已生成广州内部账单: {len(df_gz)} 行")
            except Exception as e:
                st.error(f"处理账单失败: {e}")
            if st.session_state.get('shanghai_files'):
                try:
                    shanghai_dfs = []
                    for f in st.session_state['shanghai_files']:
                        df = pd.read_excel(f)
                        shanghai_dfs.append(df)
                        st.write(f"  - {f.name}: {len(df)} 行")
                    if shanghai_dfs:
                        df_sh_merged = pd.concat(shanghai_dfs, ignore_index=True)
                        st.write(f"合并后总行数: {len(df_sh_merged)}")
                        results['shanghai_merged'] = df_sh_merged
                        sh_buffer = BytesIO()
                        with pd.ExcelWriter(sh_buffer, engine='openpyxl') as writer:
                            df_sh_merged.to_excel(writer, sheet_name='上海合并数据', index=False)
                        sh_buffer.seek(0)
                        processed_files['上海合并数据表.xlsx'] = sh_buffer.getvalue()
                        log_audit("系统", "ANNUAL", "shanghai_merge", "", f"合并上海数据 {len(df_sh_merged)} 行")
                        st.success(f"✅ 已生成上海合并数据表: {len(df_sh_merged)} 行")
                except Exception as e:
                    st.error(f"合并上海数据失败: {e}")
            else:
                st.info("未上传上海社保数据，跳过合并步骤")
            try:
                all_data = []
                if 'gz_bill' in results:
                    all_data.append(results['gz_bill'])
                if 'shanghai_merged' in results:
                    all_data.append(results['shanghai_merged'])
                if st.session_state.get('to_client_file'):
                    df_client = pd.read_excel(st.session_state['to_client_file'])
                else:
                    df_client = pd.DataFrame()
                if all_data:
                    df_merged = pd.concat(all_data, ignore_index=True)
                    results['final_data'] = df_merged
                    st.markdown("**📊 计算项1：单位缴费基数**")
                    base_col = None
                    for col in df_merged.columns:
                        if '基数' in col or '基数总额' in col:
                            base_col = col; break
                    if base_col:
                        total_base = df_merged[base_col].sum()
                        st.metric("单位缴费基数（全年合计）", f"{total_base:,.2f}")
                        results['total_base'] = total_base
                    else:
                        st.warning("未找到基数列，请确认数据格式")
                        results['total_base'] = 0
                    st.markdown("**📊 计算项2：本期实缴金额**")
                    amount_cols = []
                    for col in df_merged.columns:
                        if any(kw in col for kw in ['实缴', '缴费', '金额']):
                            amount_cols.append(col)
                    if amount_cols:
                        total_amount = 0
                        for col in amount_cols:
                            total_amount += df_merged[col].sum()
                        st.metric("本期实缴金额（各险种合计）", f"{total_amount:,.2f}")
                        results['total_amount'] = total_amount
                    else:
                        st.warning("未找到金额列，请确认数据格式")
                        results['total_amount'] = 0
                    client_buffer = BytesIO()
                    with pd.ExcelWriter(client_buffer, engine='openpyxl') as writer:
                        df_merged.to_excel(writer, sheet_name='年审汇总数据', index=False)
                        summary_df = pd.DataFrame([
                            ['统计项', '金额'],
                            ['单位缴费基数（全年）', results.get('total_base', 0)],
                            ['本期实缴金额（合计）', results.get('total_amount', 0)],
                            ['数据来源', '系统账单 + 上海社保数据'],
                            ['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
                        ])
                        summary_df.to_excel(writer, sheet_name='年审汇总', index=False, header=False)
                    client_buffer.seek(0)
                    processed_files['to客户_年审交付文件.xlsx'] = client_buffer.getvalue()
                    log_audit("系统", "ANNUAL", "to_client", "", f"生成to客户文件 {len(df_merged)} 行")
                    st.success(f"✅ 已生成to客户交付文件: {len(df_merged)} 行")
                else:
                    st.warning("请上传to客户模板或确认数据完整")
            except Exception as e:
                st.error(f"生成交付文件失败: {e}")
            if processed_files:
                st.markdown("---")
                st.markdown("### 📦 步骤3：归档打包")
                st.write(f"共 {len(processed_files)} 份文件待归档")
                for fname in processed_files.keys():
                    st.write(f"  📄 {fname}")
                zip_buffer = BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w') as zf:
                    for fname, data in processed_files.items():
                        zf.writestr(fname, data)
                zip_buffer.seek(0)
                st.download_button("📥 下载全部归档文件（ZIP）", data=zip_buffer, file_name=f"年审归档_{datetime.now().strftime('%Y%m%d_%H%M')}.zip", mime="application/zip")
                st.markdown("**单独下载**")
                cols = st.columns(min(len(processed_files), 4))
                for idx, (fname, data) in enumerate(processed_files.items()):
                    with cols[idx % 4]:
                        st.download_button(f"📄 {fname[:15]}...", data=BytesIO(data), file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.warning("未生成任何文件，请检查数据源")
    if 'results' in st.session_state:
        st.markdown("---")
        st.subheader("📊 处理结果汇总")
        results = st.session_state['results']
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("广州员工数", len(results.get('gz_bill', [])) if 'gz_bill' in results else 0)
        with col2:
            st.metric("单位缴费基数", f"{results.get('total_base', 0):,.0f}")
        with col3:
            st.metric("本期实缴金额", f"{results.get('total_amount', 0):,.0f}")

elif page == "🔐 审计日志":
    st.subheader("🔐 审计日志")
    st.markdown("**所有系统操作记录**")
    logs = load_audit_log(500)
    if logs:
        df_logs = pd.DataFrame(logs)
        st.dataframe(df_logs[['created_at', 'user_name', 'action_type', 'target_type', 'target_id', 'details']], use_container_width=True)
        st.markdown("**操作统计**")
        action_counts = df_logs['action_type'].value_counts()
        st.dataframe(pd.DataFrame({'操作类型': action_counts.index, '次数': action_counts.values}), use_container_width=True)
        search_term = st.text_input("搜索日志")
        if search_term:
            filtered = df_logs[df_logs['details'].str.contains(search_term, case=False, na=False) | 
                              df_logs['action_type'].str.contains(search_term, case=False, na=False)]
            st.dataframe(filtered, use_container_width=True)
    else:
        st.info("暂无操作日志")

elif page == "📧 邮件发送配置":
    show_email_config()

elif page == "📊 数据版本对比":
    show_data_versions()

# ===== 底部：快速生成报表 =====
st.markdown("---")
st.subheader("🚀 快速生成报表")

companies = load_companies()
if not companies:
    st.info("👈 请先在「数据导入」页面上传包含公司/城市数据的Excel")
else:
    valid_companies = [c for c in companies if c['province']]
    if len(valid_companies) < len(companies):
        st.warning(f"⚠️ 有 {len(companies) - len(valid_companies)} 家公司的省份无法识别，将使用全局默认规则")
        companies = valid_companies
        if not companies:
            st.stop()
    all_provinces = sorted(set(c['province'] for c in companies if c['province']))
    
    st.markdown("**批量选择公司（按省份/城市）**")
    batch_province = st.selectbox("批量选择省份", [""] + all_provinces, key="batch_province")
    batch_cities = []
    if batch_province:
        batch_cities = sorted(set(c['city'] for c in companies if c['province'] == batch_province))
    batch_city = st.selectbox("批量选择城市（可选）", [""] + batch_cities, key="batch_city")
    if st.button("📌 选中该地区所有公司", key="batch_select_companies"):
        if batch_province:
            selected = [c['company_name'] for c in companies if c['province'] == batch_province and (not batch_city or c['city'] == batch_city)]
            st.session_state['batch_selected_companies'] = selected
            st.success(f"已选中 {len(selected)} 家公司")
            st.rerun()
    
    col1, col2, col3 = st.columns(3)
    with col1:
        province = st.selectbox("省份", [""] + all_provinces, key="report_province")
        if province:
            cities = sorted(set(c['city'] for c in companies if c['province'] == province))
        else:
            cities = sorted(set(c['city'] for c in companies))
        city = st.selectbox("城市", [""] + cities, key="report_city")
    with col2:
        if province and city:
            districts = sorted(set(c['district'] for c in companies if c['province'] == province and c['city'] == city))
        else:
            districts = []
        district = st.selectbox("区县", [""] + districts, key="report_district")
        if province and city:
            company_list = [c for c in companies if c['province'] == province and c['city'] == city and (not district or c['district'] == district)]
        else:
            company_list = []
        company_names = [c['company_name'] for c in company_list]
        # ✅ 修复：过滤掉不在当前 company_names 中的默认值
        default_selected = st.session_state.get('batch_selected_companies', [])
        valid_default = [name for name in default_selected if name in company_names]
        selected_company_names = st.multiselect("公司（可多选）", company_names, default=valid_default, key="report_companies")
    with col3:
        report_type = st.selectbox("报表类型", ["", "增值税", "社保", "公积金", "个人所得税", "企业所得税", "年度汇算清缴"], key="report_type")
        period_type = st.selectbox("统计口径", ["月度（固定月份）", "累计（1-12月）", "自定义月份范围"], key="period_type")
        if period_type == "月度（固定月份）":
            month = st.selectbox("月份", list(range(1,13)), index=11, key="report_month")
            period_label = f"月度（{month}月）"
            custom_period = f"{month}月"
        elif period_type == "累计（1-12月）":
            period_label = "累计（1-12月）"
            custom_period = "1-12月"
        else:
            start_month = st.selectbox("起始月份", list(range(1,13)), index=0, key="start_month")
            end_month = st.selectbox("结束月份", list(range(1,13)), index=11, key="end_month")
            if start_month <= end_month:
                period_label = f"自定义（{start_month}月-{end_month}月）"
                custom_period = f"{start_month}月-{end_month}月"
            else:
                st.error("起始月份不能大于结束月份")
                period_label = "自定义"
                custom_period = ""
    selected_companies = [c for c in company_list if c['company_name'] in selected_company_names]
    if selected_companies and report_type:
        matched, match_level, candidates = match_template_with_details(province, city, district, report_type)
        custom_templates = load_custom_templates()
        all_templates = load_templates()
        recommended, reason = recommend_template_with_score(province, city, report_type, all_templates)
        if recommended:
            st.info(f"💡 推荐模板：**{recommended['template_name']}**（{reason}）")
        else:
            st.info(f"💡 {reason}，请从下方选择")
        options = {}
        if matched:
            options[f"✅ 官方模板：{matched['template_name']}（{match_level}）"] = {'type': 'official', 'data': matched}
        for c in candidates:
            if c['id'] != (matched['id'] if matched else ''):
                options[f"📄 官方模板：{c['template_name']}（{c['province']}）"] = {'type': 'official', 'data': c}
        for ct in custom_templates:
            options[f"⭐ 自定义模板：{ct['name']}"] = {'type': 'custom', 'data': ct}
        options["🔄 通用模板（系统内置）"] = {'type': 'general', 'data': None}
        if options:
            default_idx = 0
            keys = list(options.keys())
            if recommended:
                for i, k in enumerate(keys):
                    if recommended['template_name'] in k:
                        default_idx = i
                        break
            elif matched:
                for i, k in enumerate(keys):
                    if "✅" in k:
                        default_idx = i
                        break
            selected_key = st.selectbox("选择模板", keys, index=default_idx, key="template_choice")
            template_choice = options[selected_key]
        else:
            template_choice = {'type': 'general', 'data': None}
        selected_template = None
        if template_choice['type'] == 'official':
            selected_template = template_choice['data']
        elif template_choice['type'] == 'custom':
            selected_template = template_choice['data']
        else:
            selected_template = {
                'id': 'gen001',
                'template_name': f'{report_type}通用申报表',
                'template_version': 'v1.0',
                'source_authority': '系统内置',
                'publish_date': datetime.now().strftime('%Y-%m-%d'),
                'required_fields': '纳税人识别号,公司名称,申报金额',
                'source_url': '#'
            }
        with st.expander("📋 依据匹配（来源 / 模板 / 规则）", expanded=True):
            st.markdown("**匹配结果**")
            col_a, col_b = st.columns(2)
            with col_a:
                st.markdown("**来源**")
                if template_choice['type'] == 'official':
                    st.write(f"来源名称：{selected_template.get('source_authority', '未知')}")
                    st.write(f"来源URL：{selected_template.get('source_url', '#')}")
                    st.write(f"文档版本：{selected_template.get('template_version', 'v1.0')}")
                elif template_choice['type'] == 'custom':
                    st.write("来源：自定义模板")
                    st.write(f"名称：{selected_template['name']}")
                else:
                    st.write("来源：系统内置通用模板")
                st.markdown("**模板**")
                st.write(f"模板名称：{selected_template['template_name']}")
                st.write(f"模板版本：{selected_template.get('template_version', 'v1.0')}")
                st.write(f"匹配级别：{match_level if matched else '通用模板'}")
                st.write(f"使用次数：{selected_template.get('use_count', 0)} 次")
            with col_b:
                st.markdown("**规则**")
                if selected_companies:
                    first_comp = selected_companies[0]
                    rule = get_rule_for_city(first_comp['city'], first_comp.get('province'))
                    if rule:
                        st.write(f"规则来源：{rule.get('source_quote', '未配置')}")
                        st.write(f"规则版本：{rule.get('rule_version', 'v1.0')}")
                        st.write(f"来源链接：{rule.get('source_url', '#')}")
                        st.write(f"来源标题：{rule.get('source_title', '')}")
                        st.write(f"生效日期：{rule.get('effective_date', '未知')}")
                        st.write(f"使用次数：{rule.get('use_count', 0)} 次")
                    else:
                        st.write("规则：未匹配，将使用默认值")
                else:
                    st.write("规则：待选择公司后显示")
            st.markdown("**所有公司规则匹配状态**")
            rule_status = []
            for comp in selected_companies:
                r = get_rule_for_city(comp['city'], comp.get('province'))
                if r:
                    rule_status.append(f"{comp['company_name']} → {comp['city']} (规则：{r.get('source_quote', '默认')}, 版本：{r.get('rule_version', 'v1.0')})")
                else:
                    rule_status.append(f"{comp['company_name']} → {comp['city']} (⚠️ 将使用默认值)")
            st.write("\n".join(rule_status))
        with st.expander("📋 字段映射预览", expanded=False):
            fields = selected_template.get('required_fields', '').split(',')
            if fields and fields[0]:
                st.markdown("**字段列表**")
                mapping_data = []
                for f in fields:
                    source = selected_template.get('field_mapping_source', '自动映射')
                    mapping_data.append({'字段': f, '来源字段': f, '映射方式': source})
                st.dataframe(pd.DataFrame(mapping_data), use_container_width=True)
        st.warning("⚠️ 根据合规要求，所有正式导出必须经过人工复核。")
        reviewed = st.checkbox("✅ 我已人工复核确认数据无误，并核实了官方来源", value=False, key="final_review")
        export_type = st.radio("导出类型", ["验证版（可跳过复核）", "正式版（需完成复核）"], key="export_type_radio", horizontal=True)
        if export_type == "正式版（需完成复核）" and not reviewed:
            st.error("导出正式版必须勾选复核确认")
            disabled = True
        else:
            disabled = not reviewed
        if st.button("📥 生成报表", disabled=disabled, key="generate_report"):
            if not selected_template:
                st.error("请先选择模板")
            else:
                if 'validation_report' in st.session_state and st.session_state['validation_report']['error_rows'] > 0:
                    st.warning("⚠️ 当前数据存在异常（见数据质量报告），是否继续生成？")
                    if not st.checkbox("继续生成（忽略异常）", key="ignore_errors"):
                        st.stop()
                batch_id = generate_batch_id()
                batch_name = f"批量导出_{datetime.now().strftime('%Y%m%d_%H%M')}"
                save_batch_job({
                    'id': batch_id,
                    'batch_name': batch_name,
                    'created_at': datetime.now().isoformat(),
                    'status': 'processing',
                    'total_companies': len(selected_companies),
                    'total_reports': len(selected_companies),
                    'review_status': 'pending',
                    'parameters': {'province': province, 'city': city, 'report_type': report_type, 'period': period_label},
                    'created_by': '系统'
                })
                save_verification_status({
                    'batch_id': batch_id,
                    'source_verified': 0,
                    'template_verified': 0,
                    'rule_verified': 0,
                    'data_verified': 0,
                    'reviewer_name': '',
                    'verified_at': datetime.now().isoformat(),
                    'export_type': '验证版' if export_type == "验证版（可跳过复核）" else '正式版'
                })
                generated_files = []
                summary = []
                errors = []
                job_details = []
                data_source_text = st.session_state.get('data_sheet_name', '未知')
                progress_bar = st.progress(0)
                status_text = st.empty()
                total = len(selected_companies)
                increment_template_use(selected_template['id'])
                for idx, comp in enumerate(selected_companies):
                    status_text.text(f"正在处理 {idx+1}/{total}: {comp['company_name']}")
                    progress_bar.progress((idx + 1) / total)
                    try:
                        rule = get_rule_for_city(comp['city'], comp.get('province'))
                        if rule and rule.get('id'):
                            increment_rule_use(rule['id'])
                        if rule is None:
                            rule = {'unit_social': 0.16, 'personal_social': 0.08, 'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 0, 'social_max': 999999, 'fund_min': 0, 'fund_max': 999999, 'source_quote': '全局默认', 'rule_version': 'v1.0', 'source_title': '系统内置默认值', 'source_url': '#', 'source_publish_date': datetime.now().strftime('%Y-%m-%d'), 'applicable_region': '全国', 'official_channel': '系统内置'}
                        fields = selected_template.get('required_fields', '').split(',')
                        if not fields or not fields[0]:
                            fields = ['纳税人识别号', '公司名称', '申报金额']
                        if 'imported_df' in st.session_state and st.session_state['imported_df'] is not None:
                            df_data = st.session_state['imported_df']
                            company_col = None
                            for col in df_data.columns:
                                if any(kw in col.lower() for kw in COMPANY_KEYWORDS):
                                    company_col = col; break
                            if company_col:
                                df_comp = df_data[df_data[company_col] == comp['company_name']]
                                if not df_comp.empty:
                                    row_data = []
                                    for f in fields:
                                        matched_col = None
                                        for col in df_data.columns:
                                            if f in str(col) or str(col) in f:
                                                matched_col = col; break
                                        if matched_col:
                                            row_data.append(df_comp.iloc[0][matched_col])
                                        else:
                                            row_data.append('')
                                else:
                                    row_data = [''] * len(fields)
                            else:
                                row_data = [''] * len(fields)
                        else:
                            sample_data = {
                                '纳税人识别号': comp.get('tax_id', ''),
                                '公司名称': comp['company_name'],
                                '销售额': '100,000.00',
                                '进项税额': '13,000.00',
                                '应纳税额': '0.00',
                                '单位名称': comp['company_name'],
                                '社保登记号': 'SH123456',
                                '基数': '8,000.00',
                                '单位金额': str(round(8000 * rule['unit_social'], 2)) if rule else '1,280.00',
                                '个人金额': str(round(8000 * rule['personal_social'], 2)) if rule else '640.00',
                                '单位比例': str(round(rule['unit_fund'] * 100, 1)) if rule else '12.0',
                                '个人比例': str(round(rule['personal_fund'] * 100, 1)) if rule else '12.0',
                                '公积金账号': 'GJJ123456',
                                '收入额': '100,000.00',
                                '专项扣除': '0.00',
                                '营业收入': '1,000,000.00',
                                '营业成本': '600,000.00',
                                '应纳税所得额': '100,000.00',
                                '申报金额': '100,000.00',
                                '全年收入': '12,000,000.00',
                                '全年成本': '7,200,000.00',
                                '已预缴税额': '150,000.00',
                                '应补退税额': '0.00'
                            }
                            row_data = [sample_data.get(f, '') for f in fields]
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "申报表"
                        ws.append(fields)
                        ws.append(row_data)
                        ws.insert_rows(1)
                        ws['A1'] = f'【系统生成 - {export_type}】统计口径：{period_label}'
                        ws['A1'].font = Font(color='FF0000', bold=True, size=14)
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fields) if fields else 1)
                        ws['A1'].alignment = Alignment(horizontal='center')
                        ws['A1'].fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')
                        ws.insert_rows(2)
                        ws['A2'] = f'模板名称：{selected_template["template_name"]}  版本：{selected_template.get("template_version", "v1.0")}  使用次数：{selected_template.get("use_count", 0)}'
                        ws['A2'].font = Font(color='666666', size=10)
                        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(fields) if fields else 1)
                        ws.insert_rows(3)
                        ws['A3'] = f'来源：{selected_template.get("source_authority", "系统内置")}  发布日期：{selected_template.get("publish_date", datetime.now().strftime("%Y-%m-%d"))}'
                        ws['A3'].font = Font(color='666666', size=10)
                        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(fields) if fields else 1)
                        ws.insert_rows(4)
                        ws['A4'] = f'数据来源：{data_source_text}  统计口径：{period_label}'
                        ws['A4'].font = Font(color='666666', size=10)
                        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(fields) if fields else 1)
                        ws.insert_rows(5)
                        ws['A5'] = f'规则来源：{rule.get("source_quote", "未配置")}  规则版本：{rule.get("rule_version", "v1.0")}  使用次数：{rule.get("use_count", 0)}'
                        ws['A5'].font = Font(color='666666', size=10)
                        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(fields) if fields else 1)
                        ws.insert_rows(6)
                        ws['A6'] = f'来源链接：{rule.get("source_url", "#")}  来源标题：{rule.get("source_title", "")}'
                        ws['A6'].font = Font(color='666666', size=10)
                        ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(fields) if fields else 1)
                        ws_annual = wb.create_sheet("年检汇总")
                        ws_annual.append(['年检汇总数据'])
                        ws_annual.merge_cells('A1:B1')
                        ws_annual['A1'].font = Font(bold=True, size=12)
                        if 'imported_df' in st.session_state and st.session_state['imported_df'] is not None:
                            df_all = st.session_state['imported_df']
                            social_col = None
                            fund_col = None
                            unit_col = None
                            personal_col = None
                            total_col = None
                            people_col = None
                            for col in df_all.columns:
                                col_str = str(col)
                                if '社保' in col_str and '合计' in col_str:
                                    if '单位' in col_str:
                                        social_col = col
                                    elif '个人' in col_str:
                                        personal_col = col
                                elif '公积金' in col_str and '合计' in col_str:
                                    fund_col = col
                                elif '单位总费用' in col_str:
                                    unit_col = col
                                elif '个人总费用' in col_str:
                                    personal_col = col
                                elif '全部总费用' in col_str or '总金额' in col_str:
                                    total_col = col
                                elif '参保人数' in col_str:
                                    people_col = col
                            if social_col:
                                social_total = df_all[social_col].sum()
                            else:
                                social_total = 0
                            if fund_col:
                                fund_total = df_all[fund_col].sum()
                            else:
                                fund_total = 0
                            if people_col:
                                total_people = df_all[people_col].sum()
                            else:
                                total_people = len(df_all)
                            if unit_col and personal_col:
                                unit_total = df_all[unit_col].sum()
                                personal_total = df_all[personal_col].sum()
                                grand_total = df_all[total_col].sum() if total_col else (unit_total + personal_total)
                            else:
                                unit_total = social_total
                                personal_total = personal_col if personal_col else 0
                                grand_total = social_total + fund_total if fund_total else social_total
                        else:
                            total_people = 0
                            social_total = 0
                            fund_total = 0
                            unit_total = 0
                            personal_total = 0
                            grand_total = 0
                        ws_annual.append(['公司名称', comp['company_name']])
                        ws_annual.append(['所属城市', comp['city']])
                        ws_annual.append(['统计口径', period_label])
                        ws_annual.append(['参保人数（全年）', int(total_people) if total_people else 0])
                        ws_annual.append(['全年社保缴费基数总额', round(social_total, 2)])
                        ws_annual.append(['全年公积金缴费基数总额', round(fund_total, 2)])
                        ws_annual.append(['单位全年缴费总额', round(unit_total, 2)])
                        ws_annual.append(['个人全年缴费总额', round(personal_total, 2)])
                        ws_annual.append(['全年总费用', round(grand_total, 2)])
                        ws_annual.append(['报告生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                        ws_annual.append(['数据来源', data_source_text])
                        ws_annual.append(['规则来源', rule.get('source_quote', '未配置')])
                        ws_annual.append(['规则版本', rule.get('rule_version', 'v1.0')])
                        ws_annual.append(['来源链接', rule.get('source_url', '#')])
                        ws_annual.append(['来源标题', rule.get('source_title', '')])
                        audit = wb.create_sheet("审计日志")
                        audit.append(['操作时间', '操作类型', '操作人', '详情'])
                        audit.append([datetime.now().isoformat(), 'GENERATED', '系统', f'公司:{comp["company_name"]}, 城市:{comp["city"]}, 模板:{selected_template["template_name"]}, 统计口径:{period_label}, 规则:{rule.get("source_quote", "未配置")}'])
                        output = BytesIO()
                        wb.save(output)
                        output.seek(0)
                        fname = f"{comp['company_name']}_{report_type}_{period_label.replace('（','_').replace('）','').replace('-','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                        generated_files.append((fname, output.getvalue()))
                        summary.append({'公司': comp['company_name'], '城市': comp['city'], '模板': selected_template['template_name'], '统计口径': period_label, '规则来源': rule.get('source_quote', '未配置'), '规则版本': rule.get('rule_version', 'v1.0'), '状态': '待复核'})
                        job_details.append({'id': str(uuid.uuid4())[:8], 'batch_id': batch_id, 'company_id': comp['id'], 'company_name': comp['company_name'], 'city': comp['city'], 'province': comp.get('province', ''), 'report_type': report_type, 'period_type': period_label, 'status': 'success', 'error_message': '', 'file_name': fname, 'file_data': output.getvalue(), 'generated_at': datetime.now().isoformat(), 'rule_source': rule.get('source_quote', '未配置'), 'data_source': data_source_text})
                        save_export({
                            'id': str(uuid.uuid4())[:8],
                            'company_id': comp['id'],
                            'template_id': selected_template.get('id', 'gen001'),
                            'company_name': comp['company_name'],
                            'city': comp['city'],
                            'province': comp.get('province', ''),
                            'report_type': report_type,
                            'period_type': period_label,
                            'generated_at': datetime.now().isoformat(),
                            'review_status': 'pending',
                            'file_name': fname,
                            'file_data': output.getvalue(),
                            'data_source': data_source_text,
                            'month_used': period_label,
                            'year_used': datetime.now().strftime('%Y'),
                            'custom_period': custom_period if period_type == "自定义月份范围" else '',
                            'batch_id': batch_id,
                            'job_name': batch_name,
                            'field_mapping': json.dumps({f: f for f in fields})
                        })
                    except Exception as e:
                        errors.append(f"{comp['company_name']}: {str(e)}")
                        job_details.append({'id': str(uuid.uuid4())[:8], 'batch_id': batch_id, 'company_id': comp.get('id', ''), 'company_name': comp['company_name'], 'city': comp.get('city', ''), 'province': comp.get('province', ''), 'report_type': report_type, 'period_type': period_label, 'status': 'error', 'error_message': str(e), 'file_name': '', 'file_data': None, 'generated_at': datetime.now().isoformat(), 'rule_source': '', 'data_source': data_source_text})
                progress_bar.empty()
                status_text.empty()
                save_job_details(job_details)
                update_batch_status(batch_id, 'completed', 'pending')
                log_audit("系统", "GENERATE", "batch", batch_id, f"生成报表批次 {batch_name}，{len(generated_files)} 份")
                if errors:
                    for err in errors:
                        st.warning(err)
                if generated_files:
                    st.success(f"✅ 成功生成 {len(generated_files)} 份报表（批次ID：{batch_id}，状态：已完成）")
                    st.dataframe(pd.DataFrame(summary), use_container_width=True)
                    if PDF_AVAILABLE:
                        try:
                            df_summary = pd.DataFrame(summary)
                            html_content = df_to_html_table(df_summary)
                            full_html = create_pdf_from_html(html_content, f"报表_{datetime.now().strftime('%Y%m%d')}")
                            pdf_buffer = BytesIO()
                            pdfkit.from_string(full_html, pdf_buffer)
                            pdf_buffer.seek(0)
                            st.download_button("📄 下载报表摘要（PDF）", data=pdf_buffer, file_name=f"报表摘要_{datetime.now().strftime('%Y%m%d')}.pdf", mime="application/pdf")
                        except Exception as e:
                            st.warning(f"PDF导出失败（请确保已安装wkhtmltopdf）：{e}")
                    else:
                        st.warning("💡 PDF导出需要安装 pdfkit 和 wkhtmltopdf")
                    if len(generated_files) > 1:
                        zip_buffer = BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w') as zf:
                            for fname, data in generated_files:
                                zf.writestr(fname, data)
                        zip_buffer.seek(0)
                        st.download_button("📦 下载全部报表（ZIP）", data=zip_buffer, file_name=f"报表_{datetime.now().strftime('%Y%m%d')}.zip", mime="application/zip")
                    else:
                        fname, data = generated_files[0]
                        st.download_button(f"📥 下载 {fname}", data=BytesIO(data), file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
